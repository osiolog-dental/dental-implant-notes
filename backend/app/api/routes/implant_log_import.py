from __future__ import annotations

import io
import re
import uuid
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.db.session import get_db
from app.models.abutment import Abutment
from app.models.clinic import Clinic
from app.models.fpd import ProstheticFPD
from app.models.implant import Implant
from app.models.overdenture import Overdenture
from app.models.patient import Patient
from app.models.user import User

router = APIRouter(tags=["implant-log-import"])

# ── Downloadable bulk-import template ────────────────────────────────────────
# One row = one implant. Only the first row of a patient's visit needs Patient
# Name / Contact / Clinic / Date / Total No. of Implants filled in — later rows
# for the same visit (same date, additional teeth) are left blank in those
# columns. Column positions here MUST match the parsing logic once it's
# updated to read this layout (tracked separately from template generation).
_TEMPLATE_HEADERS = [
    "S.no", "Patient Name", "Contact Number", "Clinic Name", "Clinic Address",
    "Date of Implant Placement", "Total No. of Implants (this visit)",
    "Tooth Site", "Width (mm)", "Length (mm)", "Implant Cover", "Brand",
    "Product Line", "Second Stage Date", "Abutment Type", "Date of Crown Fixed",
    "Prosthetic Type", "Crown Type", "Status",
]
_TEMPLATE_COL_WIDTHS = [6, 22, 15, 20, 22, 16, 14, 10, 10, 10, 16, 16, 14, 16, 20, 16, 14, 16, 12]
_TEMPLATE_DATA_START_ROW = 4
_TEMPLATE_DATA_ROWS = 300
_COVER_OPTIONS = ["Healing Abutment", "Cover Screw"]
_CROWN_OPTIONS = ["Zirconia Crown", "PFM Crown", "Metal Crown", "Composite Crown", "Provisional Crown"]
_PROSTHETIC_OPTIONS = [
    "Single Crown",
    *[f"{n} unit FPD" for n in range(2, 13)],
    "Hybrid Denture", "Malo Bridge", "Overdenture",
]
_STATUS_OPTIONS = ["Pending", "Completed"]

_BRAND_TEAL = "82A098"
_LIGHT_TEAL_FILL = PatternFill(start_color="EAF3F1", end_color="EAF3F1", fill_type="solid")
_HEADER_FILL = PatternFill(start_color=_BRAND_TEAL, end_color=_BRAND_TEAL, fill_type="solid")


def _build_template_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Patient & Implant Log"

    ws["A1"] = "OSIOLOG — Bulk Implant Log"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    header_row = 3
    for col_idx, header in enumerate(_TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for col_idx, width in enumerate(_TEMPLATE_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 30
    ws.freeze_panes = f"A{_TEMPLATE_DATA_START_ROW}"

    last_data_row = _TEMPLATE_DATA_START_ROW + _TEMPLATE_DATA_ROWS - 1

    # Hidden helper columns (T, U) that carry each row's "block start row" and
    # "implants in this block" down from the nearest patient row above it —
    # this is what lets the conditional formatting formula below highlight
    # exactly N rows for a block of N implants without a fixed row count.
    ws["T2"] = 0
    ws["U2"] = 0
    for row in range(_TEMPLATE_DATA_START_ROW, last_data_row + 1):
        prev = row - 1
        ws[f"T{row}"] = f'=IF(B{row}<>"",ROW(),T{prev})'
        ws[f"U{row}"] = f'=IF(B{row}<>"",G{row},U{prev})'
    ws.column_dimensions["T"].hidden = True
    ws.column_dimensions["U"].hidden = True

    data_range = f"A{_TEMPLATE_DATA_START_ROW}:S{last_data_row}"
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=[f"ROW()<($T{_TEMPLATE_DATA_START_ROW}+$U{_TEMPLATE_DATA_START_ROW})"], fill=_LIGHT_TEAL_FILL),
    )

    # Dropdowns
    cover_dv = DataValidation(type="list", formula1=f'"{",".join(_COVER_OPTIONS)}"', allow_blank=True)
    ws.add_data_validation(cover_dv)
    cover_dv.add(f"K{_TEMPLATE_DATA_START_ROW}:K{last_data_row}")

    crown_dv = DataValidation(type="list", formula1=f'"{",".join(_CROWN_OPTIONS)}"', allow_blank=True)
    ws.add_data_validation(crown_dv)
    crown_dv.add(f"R{_TEMPLATE_DATA_START_ROW}:R{last_data_row}")

    status_dv = DataValidation(type="list", formula1=f'"{",".join(_STATUS_OPTIONS)}"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f"S{_TEMPLATE_DATA_START_ROW}:S{last_data_row}")

    # Prosthetic type list is long enough to risk the 255-char inline-list
    # limit, so it lives on its own sheet and the dropdown references that range.
    lists_ws = wb.create_sheet("Lists")
    lists_ws["A1"] = "Prosthetic Type Options"
    for i, option in enumerate(_PROSTHETIC_OPTIONS, start=2):
        lists_ws[f"A{i}"] = option
    lists_ws.sheet_state = "hidden"
    prosthetic_dv = DataValidation(
        type="list",
        formula1=f"Lists!$A$2:$A${1 + len(_PROSTHETIC_OPTIONS)}",
        allow_blank=True,
    )
    ws.add_data_validation(prosthetic_dv)
    prosthetic_dv.add(f"Q{_TEMPLATE_DATA_START_ROW}:Q{last_data_row}")

    # Instructions sheet
    instr_ws = wb.create_sheet("Instructions")
    instr_ws.column_dimensions["A"].width = 90
    instructions = [
        ("OSIOLOG — Bulk Implant Log — Instructions", True),
        ("", False),
        ("ONE ROW = ONE IMPLANT.", True),
        ("", False),
        ("For a patient's visit, fill in Patient Name, Contact Number, Clinic Name, Clinic Address,", False),
        ("Date of Implant Placement, and Total No. of Implants ONLY on the FIRST row of that visit.", False),
        ("If more than one implant was placed that same visit, leave those same columns BLANK on the", False),
        ("rows below it — one row per additional tooth. Fill in Tooth Site, Width, Length, Implant Cover,", False),
        ("Brand, Product Line, Abutment Type, Prosthetic Type, and Crown Type on every row (each row is", False),
        ("its own implant, even within the same visit).", False),
        ("", False),
        ("COLOUR HIGHLIGHT: once you type a number into \"Total No. of Implants\", that many rows", True),
        ("starting from that row will highlight automatically — a quick visual check that you've filled in", False),
        ("the right number of implant rows for that visit.", False),
        ("", False),
        ("If the same patient returns on a LATER date for more implants, start a new block for them", False),
        ("further down the sheet — write their name again on a fresh row.", False),
        ("", False),
        ("COLUMN NOTES", True),
        ("• Date of Implant Placement / Second Stage Date / Date of Crown Fixed: DD-MM-YYYY", False),
        ("• Tooth Site: FDI notation, e.g. 36, 14, 21", False),
        ("• Implant Cover: choose Healing Abutment or Cover Screw from the dropdown", False),
        ("• Abutment Type: free text, e.g. \"Stock Straight\", \"Stock Angled 17°\", \"Custom Abutment\"", False),
        ("• Prosthetic Type: choose Single Crown, an N-unit FPD (bridge), Hybrid Denture, Malo Bridge,", False),
        ("  or Overdenture from the dropdown — this decides what kind of record gets created for it", False),
        ("• Crown Type: choose from the dropdown, e.g. Zirconia Crown, PFM Crown", False),
        ("• Status: Pending (treatment still ongoing) or Completed (crown fixed, case finished) —", False),
        ("  this drives the osseointegration/follow-up tracking on the Dashboard", False),
        ("• Clinic Name / Clinic Address: created automatically if the clinic doesn't already exist", False),
        ("• Leave a row's Tooth Site blank to skip that row entirely", False),
    ]
    for i, (text, bold) in enumerate(instructions, start=1):
        cell = instr_ws[f"A{i}"]
        cell.value = text
        if bold:
            cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    return wb


@router.get("/implant-log-import/template")
async def download_implant_log_template(
    current_user: User = Depends(get_current_user),
) -> StreamingResponse:
    wb = _build_template_workbook()
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=OSIOLOG_Bulk_Implant_Log_Template.xlsx"},
    )

# Row layout must match _TEMPLATE_HEADERS / the columns generated above.
_ROW_COLUMNS = [
    "sno", "name", "contact", "clinic_name", "clinic_address", "date_placement",
    "total_implants", "tooth", "width", "length", "cover", "brand", "product_line",
    "stage2_date", "abutment_type", "crown_date", "prosthetic_type", "crown_type", "status",
]
_FPD_PATTERN = re.compile(r"^\s*(\d+)\s*unit\s*fpd\s*$", re.IGNORECASE)
_OVERDENTURE_TYPES = {"hybrid denture", "malo bridge", "overdenture"}


def _row_values(ws, row_num: int) -> dict:
    values = {}
    for idx, col in enumerate(_ROW_COLUMNS):
        v = ws.cell(row=row_num, column=idx + 1).value
        values[col] = v.strip() if isinstance(v, str) else v
    return values


def _current_stage(status, stage2_date, crown_date) -> int:
    if (status and str(status).strip().lower() == "completed") or crown_date:
        return 3
    if stage2_date:
        return 2
    return 1


def _parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str) and v.strip():
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _parse_float(v):
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


async def _get_or_create_clinic_id(db: AsyncSession, org_id: uuid.UUID, name, address=None) -> str | None:
    if not name or not str(name).strip():
        return None
    name = str(name).strip()
    address = str(address).strip() if address else None
    result = await db.execute(select(Clinic).where(Clinic.org_id == org_id, Clinic.name == name))
    clinic = result.scalar_one_or_none()
    if clinic:
        if address and not clinic.address:
            clinic.address = address
        return str(clinic.id)
    clinic = Clinic(id=uuid.uuid4(), org_id=org_id, name=name, address=address)
    db.add(clinic)
    await db.flush()
    return str(clinic.id)


@router.post("/implant-log-import")
async def import_implant_log_files(
    files: list[UploadFile] = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Accepts the bulk Patient & Implant Log template (see /implant-log-import/template):
    one row per implant, with Patient Name/Contact/Clinic/Date only filled on the first
    row of a visit and carried forward for the blank continuation rows below it.
    """
    results = []
    total_patients = 0
    total_implants = 0

    for upload in files:
        file_result = {
            "filename": upload.filename,
            "patients_created": 0,
            "implants_created": 0,
            "errors": [],
        }
        try:
            content = await upload.read()
            wb = load_workbook(io.BytesIO(content), data_only=True)
            sheet_name = "Patient & Implant Log" if "Patient & Implant Log" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet_name]

            patient = None
            clinic_id = None
            visit_date = None
            pending_group = None  # in-progress FPD/Overdenture spanning a run of rows

            async def flush_group():
                nonlocal pending_group
                if pending_group is None:
                    return
                if pending_group["kind"] == "fpd":
                    db.add(ProstheticFPD(
                        id=uuid.uuid4(),
                        case_id=None,
                        patient_id=pending_group["patient_id"],
                        tooth_numbers=pending_group["teeth"],
                        prosthetic_loading_date=pending_group["loading_date"],
                        crown_count=pending_group["crown_count"],
                        connected_implant_ids=pending_group["implant_ids"],
                        crown_type=pending_group["crown_type"],
                    ))
                else:
                    db.add(Overdenture(
                        id=uuid.uuid4(),
                        patient_id=pending_group["patient_id"],
                        tooth_numbers=pending_group["teeth"],
                        attachment_type=pending_group["prosthetic_type"],
                        connected_implant_ids=pending_group["implant_ids"],
                        prosthetic_loading_date=pending_group["loading_date"],
                        clinic_id=pending_group["clinic_id"],
                    ))
                pending_group = None

            max_row = ws.max_row or _TEMPLATE_DATA_START_ROW
            for row_num in range(_TEMPLATE_DATA_START_ROW, max_row + 1):
                vals = _row_values(ws, row_num)

                name = vals["name"]
                if name and str(name).strip():
                    await flush_group()
                    clinic_id = await _get_or_create_clinic_id(
                        db, current_user.org_id, vals["clinic_name"], vals["clinic_address"]
                    )
                    patient = Patient(
                        id=uuid.uuid4(),
                        org_id=current_user.org_id,
                        doctor_id=current_user.id,
                        name=str(name).strip(),
                        phone=str(vals["contact"]).strip() if vals["contact"] else None,
                    )
                    db.add(patient)
                    await db.flush()
                    file_result["patients_created"] += 1
                    total_patients += 1
                    visit_date = _parse_date(vals["date_placement"])

                tooth = vals["tooth"]
                if tooth in (None, ""):
                    continue
                if patient is None:
                    file_result["errors"].append(f"Row {row_num}: implant row with no patient above it — skipped")
                    continue

                try:
                    stage2_date = _parse_date(vals["stage2_date"])
                    crown_date = _parse_date(vals["crown_date"])
                    cover = str(vals["cover"]).strip().lower() if vals["cover"] else ""

                    implant = Implant(
                        id=uuid.uuid4(),
                        case_id=None,
                        patient_id=patient.id,
                        tooth_number=int(tooth),
                        brand=vals["brand"] or None,
                        implant_system=vals["product_line"] or None,
                        diameter_mm=_parse_float(vals["width"]),
                        length_mm=_parse_float(vals["length"]),
                        cover_screw=(cover == "cover screw") or None,
                        healing_abutment=(cover == "healing abutment") or None,
                        surgery_date=visit_date,
                        stage_2_date=stage2_date,
                        prosthetic_loading_date=crown_date,
                        current_stage=_current_stage(vals["status"], stage2_date, crown_date),
                        clinic_id=clinic_id,
                    )
                    db.add(implant)
                    await db.flush()
                    file_result["implants_created"] += 1
                    total_implants += 1

                    abutment_type = vals["abutment_type"]
                    if abutment_type and str(abutment_type).strip():
                        db.add(Abutment(
                            id=uuid.uuid4(),
                            patient_id=patient.id,
                            tooth_number=int(tooth),
                            abutment_type=str(abutment_type).strip(),
                            connected_implant_ids=[implant.id],
                            placement_date=stage2_date or visit_date,
                            clinic_id=clinic_id,
                        ))

                    prosthetic_type = str(vals["prosthetic_type"]).strip() if vals["prosthetic_type"] else ""
                    fpd_match = _FPD_PATTERN.match(prosthetic_type)
                    is_overdenture = prosthetic_type.lower() in _OVERDENTURE_TYPES

                    if fpd_match or is_overdenture:
                        kind = "fpd" if fpd_match else "overdenture"
                        same_run = (
                            pending_group is not None
                            and pending_group["kind"] == kind
                            and pending_group["prosthetic_type"] == prosthetic_type
                            and pending_group["patient_id"] == patient.id
                        )
                        if not same_run:
                            await flush_group()
                            pending_group = {
                                "kind": kind,
                                "patient_id": patient.id,
                                "clinic_id": clinic_id,
                                "prosthetic_type": prosthetic_type,
                                "crown_count": fpd_match.group(1) if fpd_match else None,
                                "crown_type": vals["crown_type"] or None,
                                "loading_date": crown_date,
                                "teeth": [],
                                "implant_ids": [],
                            }
                        pending_group["teeth"].append(int(tooth))
                        pending_group["implant_ids"].append(implant.id)
                        if crown_date:
                            pending_group["loading_date"] = crown_date
                    else:
                        await flush_group()
                        if prosthetic_type and prosthetic_type.lower() != "single crown":
                            file_result["errors"].append(
                                f'Row {row_num}: unrecognized Prosthetic Type "{prosthetic_type}" — implant saved without a prosthetic record'
                            )
                except (TypeError, ValueError) as exc:
                    file_result["errors"].append(f"Row {row_num} (Tooth {tooth}): {exc}")

            await flush_group()
            await db.flush()
        except Exception as exc:
            file_result["errors"].append(f"Could not read file: {exc}")

        results.append(file_result)

    await db.commit()
    return {"results": results, "totals": {"patients": total_patients, "implants": total_implants}}
