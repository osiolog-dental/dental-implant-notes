"""
Flat-URL aliases used by the PatientDetails frontend.

These mirror the nested routes but accept patient_id as a query param
or body field, and return _id aliases in the JSON for MongoDB compat.
Also includes: warranty-image upload, patient photos, edit-log, PUT for
implants and fpd-records, and patient profile-picture upload.
"""
from __future__ import annotations

import asyncio
import uuid
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import RedirectResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.db.session import get_db
from app.models.audit import AuditEvent
from app.models.case import CaseImage
from app.models.fpd import ProstheticFPD
from app.models.implant import Implant
from app.models.patient import Patient
from app.models.user import User
from app.repositories.fpd import FPDRepository
from app.repositories.implant import ImplantRepository
from app.schemas.fpd import FPDFlatCreate, FPDRead, FPDUpdate
from pydantic import BaseModel
from app.schemas.implant import ImplantFlatCreate, ImplantRead, ImplantUpdate
from app.services import s3 as s3_service

router = APIRouter(tags=["flat-routes"])


# ── /api/implants  ─────────────────────────────────────────────────────────────

@router.get("/implants", response_model=list[ImplantRead])
async def list_implants_flat(
    patient_id: uuid.UUID = Query(...),
    clinic_id: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list[ImplantRead]:
    repo = ImplantRepository(db)
    implants = await repo.list_by_patient(patient_id, current_user.org_id)
    if clinic_id:
        implants = [i for i in implants if i.clinic_id == clinic_id]
    return [ImplantRead.model_validate(i) for i in implants]


@router.post("/implants", response_model=ImplantRead, status_code=201)
async def create_implant_flat(
    body: ImplantFlatCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ImplantRead:
    data = body.model_dump()
    implant = Implant(id=uuid.uuid4(), case_id=None, **data)
    db.add(implant)
    await db.commit()
    await db.refresh(implant)
    return ImplantRead.model_validate(implant)


@router.put("/implants/{implant_id}", response_model=ImplantRead)
async def update_implant_flat(
    implant_id: uuid.UUID,
    body: ImplantUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ImplantRead:
    repo = ImplantRepository(db)
    implant = await repo.get(implant_id, current_user.org_id)
    if not implant:
        raise HTTPException(status_code=404, detail="Implant not found")
    implant = await repo.update(implant, body)
    return ImplantRead.model_validate(implant)


class StageUpdate(BaseModel):
    current_stage: int | None = None
    osseointegration_days: int | None = None
    stage_2_date: str | None = None
    stage_3_date: str | None = None


@router.patch("/implants/{implant_id}/stage", response_model=ImplantRead)
async def update_implant_stage(
    implant_id: uuid.UUID,
    body: StageUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ImplantRead:
    repo = ImplantRepository(db)
    implant = await repo.get(implant_id, current_user.org_id)
    if not implant:
        raise HTTPException(status_code=404, detail="Implant not found")
    if body.current_stage is not None:
        implant.current_stage = body.current_stage
    if body.osseointegration_days is not None:
        implant.osseointegration_days = body.osseointegration_days
    if body.stage_2_date is not None:
        from datetime import date
        implant.stage_2_date = date.fromisoformat(body.stage_2_date) if body.stage_2_date else None
    if body.stage_3_date is not None:
        from datetime import date
        implant.stage_3_date = date.fromisoformat(body.stage_3_date) if body.stage_3_date else None
    db.add(implant)
    await db.commit()
    await db.refresh(implant)
    return ImplantRead.model_validate(implant)


# ── /api/fpd-records  ──────────────────────────────────────────────────────────

@router.get("/fpd-records", response_model=list[FPDRead])
async def list_fpd_flat(
    patient_id: uuid.UUID = Query(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list[FPDRead]:
    repo = FPDRepository(db)
    records = await repo.list_by_patient(patient_id, current_user.org_id)
    return [FPDRead.model_validate(r) for r in records]


@router.post("/fpd-records", response_model=FPDRead, status_code=201)
async def create_fpd_flat(
    body: FPDFlatCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> FPDRead:
    fpd = ProstheticFPD(
        id=uuid.uuid4(),
        case_id=None,
        patient_id=body.patient_id,
        tooth_numbers=body.tooth_numbers,
        prosthetic_loading_date=body.prosthetic_loading_date,
        crown_count=body.crown_count,
        connected_implant_ids=body.connected_implant_ids,
        crown_type=body.crown_type,
        material=body.material,
        crown_material=body.crown_material,
        clinical_notes=body.clinical_notes,
        consultant_prosthodontist=body.consultant_prosthodontist,
        lab_name=body.lab_name,
    )
    db.add(fpd)
    await db.commit()
    await db.refresh(fpd)
    return FPDRead.model_validate(fpd)


@router.put("/fpd-records/{fpd_id}", response_model=FPDRead)
async def update_fpd_flat(
    fpd_id: uuid.UUID,
    body: FPDUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> FPDRead:
    repo = FPDRepository(db)
    fpd = await repo.get(fpd_id, current_user.org_id)
    if not fpd:
        raise HTTPException(status_code=404, detail="FPD record not found")
    fpd = await repo.update(fpd, body)
    return FPDRead.model_validate(fpd)


@router.patch("/fpd-records/{fpd_id}", response_model=FPDRead)
async def patch_fpd_flat(
    fpd_id: uuid.UUID,
    body: FPDUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> FPDRead:
    repo = FPDRepository(db)
    fpd = await repo.get(fpd_id, current_user.org_id)
    if not fpd:
        raise HTTPException(status_code=404, detail="FPD record not found")
    fpd = await repo.update(fpd, body)
    return FPDRead.model_validate(fpd)


@router.delete("/fpd-records/{fpd_id}")
async def delete_fpd_flat(
    fpd_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    repo = FPDRepository(db)
    fpd = await repo.get(fpd_id, current_user.org_id)
    if not fpd:
        raise HTTPException(status_code=404, detail="FPD record not found")
    await db.delete(fpd)
    await db.commit()
    return {"deleted": True}


@router.post("/fpd-records/{fpd_id}/warranty-image")
async def upload_warranty_image(
    fpd_id: uuid.UUID,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    repo = FPDRepository(db)
    fpd = await repo.get(fpd_id, current_user.org_id)
    if not fpd:
        raise HTTPException(status_code=404, detail="FPD record not found")

    content_type = file.content_type or "application/octet-stream"
    if content_type not in s3_service.ALLOWED_CONTENT_TYPES:
        raise HTTPException(status_code=400, detail=f"File type not allowed: {content_type}")

    ext = content_type.split("/")[-1].replace("jpeg", "jpg")
    s3_key = f"warranty/{current_user.org_id}/{fpd_id}.{ext}"

    image_bytes = await file.read()
    await asyncio.to_thread(s3_service.upload_object, s3_key, image_bytes, content_type)

    fpd.warranty_image_url = s3_key
    db.add(fpd)
    await db.flush()

    download_url = s3_service.generate_download_url(s3_key)
    return {"warranty_image_url": download_url}


# ── Patient photos (for PDF export)  ──────────────────────────────────────────

@router.get("/patients/{patient_id}/photos")
async def list_patient_photos(
    patient_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list[dict]:
    patient_result = await db.execute(
        select(Patient).where(
            Patient.id == patient_id,
            Patient.org_id == current_user.org_id,
            Patient.deleted_at.is_(None),
        )
    )
    patient = patient_result.scalar_one_or_none()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    from app.models.case import Case
    result = await db.execute(
        select(CaseImage)
        .join(Case, CaseImage.case_id == Case.id)
        .where(
            Case.patient_id == patient_id,
            CaseImage.status == "uploaded",
        )
        .order_by(CaseImage.uploaded_at.desc())
        .limit(20)
    )
    images = result.scalars().all()

    photos = []
    for img in images:
        try:
            url = s3_service.generate_download_url(img.s3_key)
            photos.append({"url": url, "content_type": img.content_type, "uploaded_at": img.uploaded_at.isoformat()})
        except Exception:
            continue
    return photos


# ── Patient edit log  ──────────────────────────────────────────────────────────

@router.get("/patients/{patient_id}/edit-log")
async def get_patient_edit_log(
    patient_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list[dict]:
    result = await db.execute(
        select(AuditEvent)
        .where(
            AuditEvent.org_id == current_user.org_id,
            AuditEvent.entity_id == str(patient_id),
        )
        .order_by(AuditEvent.created_at.desc())
        .limit(50)
    )
    events = result.scalars().all()
    return [
        {
            "action": e.action,
            "entity_type": e.entity_type,
            "created_at": e.created_at.isoformat(),
            "user_id": str(e.user_id),
        }
        for e in events
    ]


# ── Patient PUT (friend's frontend uses PUT, not PATCH)  ──────────────────────

@router.put("/patients/{patient_id}")
async def update_patient_put(
    patient_id: uuid.UUID,
    body: dict,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    from app.repositories.patient import PatientRepository
    from app.schemas.patient import PatientUpdate
    repo = PatientRepository(db)
    patient = await repo.get(patient_id, current_user.org_id)
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    allowed = {"name", "age", "gender", "phone", "email", "address", "medical_history",
               "alternate_email", "emergency_phone"}
    for k, v in body.items():
        if k in allowed:
            setattr(patient, k, v)
    db.add(patient)
    await db.flush()
    return {"id": str(patient.id), "name": patient.name, "_id": str(patient.id)}


# ── Patient profile picture  ──────────────────────────────────────────────────

@router.post("/patients/{patient_id}/profile-picture")
async def upload_patient_profile_picture(
    patient_id: uuid.UUID,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    patient_result = await db.execute(
        select(Patient).where(
            Patient.id == patient_id,
            Patient.org_id == current_user.org_id,
            Patient.deleted_at.is_(None),
        )
    )
    patient = patient_result.scalar_one_or_none()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    content_type = file.content_type or "image/jpeg"
    ext = content_type.split("/")[-1].replace("jpeg", "jpg")
    s3_key = f"patients/{current_user.org_id}/{patient_id}/profile.{ext}"

    image_bytes = await file.read()
    await asyncio.to_thread(s3_service.upload_object, s3_key, image_bytes, content_type)

    # Store s3_key on the patient record so /api/files/:filename can serve it
    patient.profile_picture = s3_key
    db.add(patient)
    await db.flush()

    url = s3_service.generate_download_url(s3_key)
    return {"profile_picture_url": url, "profile_picture": s3_key}


# ── Doctor (user) profile picture  ───────────────────────────────────────────

@router.post("/users/me/profile-picture")
async def upload_doctor_profile_picture(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    content_type = file.content_type or "image/jpeg"
    ext = content_type.split("/")[-1].replace("jpeg", "jpg")
    s3_key = f"doctors/{current_user.org_id}/{current_user.id}/profile.{ext}"

    image_bytes = await file.read()
    await asyncio.to_thread(s3_service.upload_object, s3_key, image_bytes, content_type)

    current_user.profile_picture_key = s3_key
    db.add(current_user)
    await db.flush()

    url = s3_service.generate_download_url(s3_key)
    return {"profile_picture_url": url}


# ── /api/files/:filename — presigned redirect for image src tags  ─────────────

@router.get("/files/{filename:path}")
async def serve_file(filename: str) -> RedirectResponse:
    """
    Redirect to a presigned S3 download URL.
    No auth required — the presigned S3 URL is self-authenticating and time-limited.
    `filename` is the S3 key stored in patient.profile_picture or fpd.warranty_image_url.
    """
    try:
        url = s3_service.generate_download_url(filename)
        return RedirectResponse(url=url, status_code=302)
    except Exception:
        raise HTTPException(status_code=404, detail="File not found")


# ── Public doctor profile (no auth required)  ─────────────────────────────────

@router.get("/public/profile/{doctor_id}")
async def get_public_profile(
    doctor_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
) -> dict:
    from app.models.case import Case
    from app.models.implant import Implant as ImplantModel
    from app.models.fpd import ProstheticFPD as FPDModel
    from app.models.clinic import Clinic

    user_result = await db.execute(
        select(User).where(User.id == doctor_id)
    )
    user = user_result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Profile not found")

    # Stats
    patients_result = await db.execute(
        select(Patient).where(Patient.org_id == user.org_id, Patient.deleted_at.is_(None))
    )
    patients = patients_result.scalars().all()
    patient_count = len(patients)

    implants_result = await db.execute(
        select(ImplantModel)
        .join(Patient, ImplantModel.patient_id == Patient.id)
        .where(Patient.org_id == user.org_id, Patient.deleted_at.is_(None))
    )
    implant_count = len(implants_result.scalars().all())

    fpd_result = await db.execute(
        select(FPDModel)
        .join(Patient, FPDModel.patient_id == Patient.id)
        .where(Patient.org_id == user.org_id, Patient.deleted_at.is_(None))
    )
    fpd_count = len(fpd_result.scalars().all())

    clinics_result = await db.execute(
        select(Clinic).where(Clinic.org_id == user.org_id)
    )
    clinics = [{"name": c.name, "address": c.address} for c in clinics_result.scalars().all()]

    profile_picture_url = None
    if user.profile_picture_key:
        try:
            profile_picture_url = s3_service.generate_download_url(user.profile_picture_key)
        except Exception:
            pass

    return {
        "id": str(user.id),
        "name": user.name,
        "email": user.email,
        "specialization": user.specialization,
        "registration_number": user.registration_number,
        "college": user.college,
        "college_place": user.college_place,
        "country": user.country,
        "bio": user.bio,
        "profile_picture": profile_picture_url,
        "clinics": clinics,
        "stats": {
            "patients": patient_count,
            "implants": implant_count,
            "fpd": fpd_count,
            "success_rate": None,
        },
    }


# ── Analytics  ────────────────────────────────────────────────────────────────

@router.get("/analytics/overview")
async def analytics_overview(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    from app.models.clinic import Clinic

    patients_result = await db.execute(
        select(Patient).where(
            Patient.org_id == current_user.org_id,
            Patient.deleted_at.is_(None),
        )
    )
    total_patients = len(patients_result.scalars().all())

    implants_result = await db.execute(
        select(Implant)
        .join(Patient, Implant.patient_id == Patient.id)
        .where(
            Patient.org_id == current_user.org_id,
            Patient.deleted_at.is_(None),
        )
    )
    all_implants = list(implants_result.scalars().all())
    total_implants = len(all_implants)

    pending_osseointegration = sum(
        1 for i in all_implants
        if i.current_stage == 1 and not i.osseointegration_success
    )

    clinics_result = await db.execute(
        select(Clinic).where(Clinic.org_id == current_user.org_id)
    )
    total_clinics = len(clinics_result.scalars().all())

    type_counts: dict[str, int] = {}
    for imp in all_implants:
        t = imp.implant_type or "Unknown"
        type_counts[t] = type_counts.get(t, 0) + 1
    implant_types = [{"_id": t, "count": c} for t, c in type_counts.items()]

    return {
        "total_patients": total_patients,
        "total_implants": total_implants,
        "pending_osseointegration": pending_osseointegration,
        "total_clinics": total_clinics,
        "implant_types": implant_types,
    }


@router.get("/analytics/financial")
async def analytics_financial(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    implants_result = await db.execute(
        select(Implant)
        .join(Patient, Implant.patient_id == Patient.id)
        .where(
            Patient.org_id == current_user.org_id,
            Patient.deleted_at.is_(None),
        )
    )
    all_implants = list(implants_result.scalars().all())
    total_implants = len(all_implants)

    RATE = {"Single": 1500, "Bridge": 4500, "Full Mouth": 25000}
    DEFAULT_RATE = 1500
    total_revenue = sum(
        RATE.get(i.implant_type or "Single", DEFAULT_RATE)
        for i in all_implants
    )
    average_per_implant = total_revenue / total_implants if total_implants else 0

    return {
        "total_revenue": total_revenue,
        "average_per_implant": round(average_per_implant, 2),
        "total_implants": total_implants,
    }


@router.get("/analytics/per-patient")
async def analytics_per_patient(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list:
    implants_result = await db.execute(
        select(Implant, Patient)
        .join(Patient, Implant.patient_id == Patient.id)
        .where(
            Patient.org_id == current_user.org_id,
            Patient.deleted_at.is_(None),
        )
    )
    rows = implants_result.all()

    RATE = {"Single": 1500, "Bridge": 4500, "Full Mouth": 25000}
    DEFAULT_RATE = 1500

    patient_map: dict[str, dict] = {}
    for implant, patient in rows:
        pid = str(patient.id)
        if pid not in patient_map:
            patient_map[pid] = {
                "patient_id": pid,
                "patient_name": patient.name,
                "total_revenue": 0,
                "implant_count": 0,
            }
        patient_map[pid]["total_revenue"] += RATE.get(implant.implant_type or "Single", DEFAULT_RATE)
        patient_map[pid]["implant_count"] += 1

    sorted_patients = sorted(patient_map.values(), key=lambda x: x["total_revenue"], reverse=True)
    return sorted_patients[:10]


@router.get("/implants/due-for-second-stage")
async def implants_due_for_second_stage(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list:
    from datetime import date as _date

    result = await db.execute(
        select(Implant, Patient)
        .join(Patient, Implant.patient_id == Patient.id)
        .where(
            Patient.org_id == current_user.org_id,
            Patient.deleted_at.is_(None),
            Implant.surgery_date.isnot(None),
            Implant.current_stage == 1,
        )
    )
    rows = result.all()

    today = _date.today()
    due = []
    for implant, patient in rows:
        days_elapsed = (today - implant.surgery_date).days
        if days_elapsed >= implant.osseointegration_days:
            due.append({
                "implant_id": str(implant.id),
                "patient_id": str(patient.id),
                "patient_name": patient.name,
                "tooth_number": implant.tooth_number,
                "brand": implant.brand,
                "case_number": None,
                "days_elapsed": days_elapsed,
                "osseointegration_days": implant.osseointegration_days,
                "surgery_date": implant.surgery_date.isoformat(),
            })

    return sorted(due, key=lambda x: x["days_elapsed"], reverse=True)


@router.get("/implants/all")
async def list_all_implants(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list[ImplantRead]:
    result = await db.execute(
        select(Implant)
        .join(Patient, Implant.patient_id == Patient.id)
        .where(
            Patient.org_id == current_user.org_id,
            Patient.deleted_at.is_(None),
        )
        .order_by(Implant.created_at.desc())
    )
    implants = result.scalars().all()
    return [ImplantRead.model_validate(i) for i in implants]


# ── Backup export / restore  ───────────────────────────────────────────────────

@router.get("/backup/export")
async def export_backup(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    from app.models.case import Case
    from app.models.implant import Implant as ImplantModel
    from app.models.fpd import ProstheticFPD as FPDModel
    from app.models.clinic import Clinic
    from datetime import timezone
    import datetime as _dt

    patients_result = await db.execute(
        select(Patient).where(Patient.org_id == current_user.org_id, Patient.deleted_at.is_(None))
    )
    patients_rows = patients_result.scalars().all()

    implants_result = await db.execute(
        select(ImplantModel)
        .join(Patient, ImplantModel.patient_id == Patient.id)
        .where(Patient.org_id == current_user.org_id, Patient.deleted_at.is_(None))
    )
    implants_rows = implants_result.scalars().all()

    fpd_result = await db.execute(
        select(FPDModel)
        .join(Patient, FPDModel.patient_id == Patient.id)
        .where(Patient.org_id == current_user.org_id, Patient.deleted_at.is_(None))
    )
    fpd_rows = fpd_result.scalars().all()

    clinics_result = await db.execute(
        select(Clinic).where(Clinic.org_id == current_user.org_id)
    )
    clinics_rows = clinics_result.scalars().all()

    def _row(obj):
        return {c.name: str(getattr(obj, c.name)) if getattr(obj, c.name) is not None else None
                for c in obj.__table__.columns}

    return {
        "version": "2.0",
        "exported_at": _dt.datetime.now(timezone.utc).isoformat(),
        "patients": [_row(p) for p in patients_rows],
        "implants": [_row(i) for i in implants_rows],
        "fpd_records": [_row(f) for f in fpd_rows],
        "clinics": [_row(c) for c in clinics_rows],
    }


@router.post("/backup/restore")
async def restore_backup(
    payload: dict,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    if payload.get("version") not in ("2.0",):
        raise HTTPException(status_code=400, detail="Unsupported backup version")

    from app.models.clinic import Clinic
    from app.models.implant import Implant as ImplantModel
    from app.models.fpd import ProstheticFPD as FPDModel

    inserted = {"patients": 0, "implants": 0, "fpd_records": 0, "clinics": 0}

    for raw in payload.get("patients", []):
        pid = uuid.UUID(raw["id"]) if raw.get("id") else uuid.uuid4()
        existing = await db.get(Patient, pid)
        if not existing:
            p = Patient(
                id=pid,
                org_id=current_user.org_id,
                doctor_id=current_user.id,
                name=raw.get("name", "Unknown"),
                age=int(raw["age"]) if raw.get("age") else None,
                gender=raw.get("gender"),
                phone=raw.get("phone"),
                email=raw.get("email"),
                address=raw.get("address"),
                medical_history=raw.get("medical_history"),
            )
            db.add(p)
            inserted["patients"] += 1

    await db.flush()
    return {"inserted": inserted}


# ── Subscription status  ───────────────────────────────────────────────────────

@router.get("/subscription/status")
async def subscription_status(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    return {
        "plan": "free",
        "used_mb": 0,
        "limit_mb": 500,
        "plan_end": None,
    }


@router.post("/subscription/upgrade")
async def subscription_upgrade(
    body: dict,
    current_user: User = Depends(get_current_user),
) -> dict:
    raise HTTPException(status_code=402, detail="Payment integration coming soon. Contact support to upgrade.")


@router.get("/ads")
async def get_ads(current_user: User = Depends(get_current_user)) -> list:
    return []


# ── Notifications  ─────────────────────────────────────────────────────────────

@router.get("/notifications/osseointegration-alerts")
async def get_osseointegration_alerts(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list:
    """Return implants within 14 days of osseointegration completion (surgery_date + 90 days)."""
    from datetime import date, timedelta
    cutoff = date.today() + timedelta(days=14)
    start = date.today()

    result = await db.execute(
        select(Implant, Patient)
        .join(Patient, Implant.patient_id == Patient.id)
        .where(
            Patient.org_id == current_user.org_id,
            Patient.deleted_at.is_(None),
            Implant.surgery_date.isnot(None),
        )
    )
    rows = result.all()

    alerts = []
    for implant, patient in rows:
        if implant.surgery_date:
            osseo_end = implant.surgery_date + timedelta(days=90)
            if start <= osseo_end <= cutoff:
                alerts.append({
                    "implant_id": implant.id,
                    "patient_id": patient.id,
                    "patient_name": patient.name,
                    "tooth_number": implant.tooth_number,
                    "brand": implant.brand,
                    "osseo_date": str(osseo_end),
                    "days_remaining": (osseo_end - date.today()).days,
                })
    return sorted(alerts, key=lambda x: x["days_remaining"])


# ── Excel export  ──────────────────────────────────────────────────────────────

@router.get("/export/implants")
async def export_implants_excel(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Download all implant records as Excel (.xlsx)."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    # Query all implants for this org
    result = await db.execute(
        select(Implant, Patient)
        .join(Patient, Implant.patient_id == Patient.id)
        .where(
            Patient.org_id == current_user.org_id,
            Patient.deleted_at.is_(None),
        )
        .order_by(Patient.name, Implant.tooth_number)
    )
    rows = result.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Implant Records"

    headers = [
        "Patient Name", "Patient Age", "Tooth #", "Implant Type",
        "Brand", "System", "Diameter (mm)", "Length (mm)",
        "Insertion Torque", "ISQ", "Connection Type",
        "Surgical Approach", "Arch", "Jaw Region",
        "Surgery Date", "Follow-up Date", "Prosthetic Loading",
        "Surgeon", "Clinic", "Outcome", "Notes"
    ]

    # Style header row
    header_fill = PatternFill("solid", fgColor="82A098")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_num, (implant, patient) in enumerate(rows, 2):
        ws.append([
            patient.name,
            patient.age,
            implant.tooth_number,
            implant.implant_type,
            implant.brand,
            implant.implant_system,
            implant.diameter_mm,
            implant.length_mm,
            implant.insertion_torque,
            implant.isq_value,
            implant.connection_type,
            implant.surgical_approach,
            implant.arch,
            implant.jaw_region,
            str(implant.surgery_date) if implant.surgery_date else "",
            str(implant.follow_up_date) if implant.follow_up_date else "",
            str(implant.prosthetic_loading_date) if implant.prosthetic_loading_date else "",
            implant.surgeon_name,
            implant.clinic_id,
            implant.implant_outcome,
            implant.notes,
        ])

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=implant-records.xlsx"},
    )
