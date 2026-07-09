from dotenv import load_dotenv
from pathlib import Path
import logging

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Configure logging early
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File, Query, Header
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
from pydantic import BaseModel, EmailStr, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta, date
from bson import ObjectId
import bcrypt
import jwt
import secrets
import uuid
import shutil

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=[os.environ.get('FRONTEND_URL', 'http://localhost:3000')],
    allow_methods=["*"],
    allow_headers=["*"],
)

JWT_ALGORITHM = "HS256"
APP_NAME = "dentalhub"

# Local file storage — files saved under backend/uploads/
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

def put_object(path: str, data: bytes, content_type: str) -> dict:
    dest = UPLOAD_DIR / path
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(data)
    return {"path": path, "size": len(data), "content_type": content_type}

def get_object(path: str) -> tuple:
    dest = UPLOAD_DIR / path
    if not dest.exists():
        raise HTTPException(status_code=404, detail="File not found")
    ext = dest.suffix.lower()
    mime_map = {
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
        ".gif": "image/gif", ".webp": "image/webp", ".pdf": "application/pdf"
    }
    content_type = mime_map.get(ext, "application/octet-stream")
    return dest.read_bytes(), content_type

# Helper Functions
def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]

# Auth Routes
@api_router.get("/")
async def root():
    return {"message": "Dental Hub API", "status": "running"}

def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password.encode("utf-8"), salt)
    return hashed.decode("utf-8")

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode("utf-8"), hashed_password.encode("utf-8"))

def create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=15),
        "type": "access"
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "refresh"
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user["_id"] = str(user["_id"])
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# Models
class DoctorRegister(BaseModel):
    email: EmailStr
    password: str
    name: str
    phone: str
    country: str
    currency: Optional[str] = 'USD'
    registration_number: str
    college: Optional[str] = None
    specialization: str
    profile_picture: Optional[str] = None
    # Extended fields
    gender: Optional[str] = None
    date_of_birth: Optional[str] = None
    address_street: Optional[str] = None
    address_city: Optional[str] = None
    address_state: Optional[str] = None
    address_zip: Optional[str] = None
    designation: Optional[str] = None
    organization: Optional[str] = None
    years_of_experience: Optional[str] = None
    primary_clinic: Optional[str] = None
    consulting_clinics: Optional[str] = None
    clinical_focus: Optional[str] = None
    education: Optional[List[Dict[str, Any]]] = []
    publications: Optional[List[Dict[str, Any]]] = []

class DoctorLogin(BaseModel):
    email: EmailStr
    password: str

class PatientCreate(BaseModel):
    name: str
    age: int
    gender: str
    phone: str
    email: Optional[str] = None
    address: Optional[str] = None
    medical_history: Optional[str] = None
    profile_picture: Optional[str] = None
    emergency_phone: Optional[str] = None
    alternate_email: Optional[str] = None

class ClinicCreate(BaseModel):
    name: str
    address: str
    phone: str
    email: Optional[str] = None

class ImplantCreate(BaseModel):
    patient_id: str
    tooth_number: int
    implant_type: str
    brand: str
    size: str
    length: str
    insertion_torque: Optional[float] = None
    connection_type: str
    surgical_approach: str
    bone_graft: Optional[str] = None
    sinus_lift_type: Optional[str] = None
    is_pterygoid: bool = False
    is_zygomatic: bool = False
    is_subperiosteal: bool = False
    notes: Optional[str] = None
    clinic_id: Optional[str] = None
    # Enhanced fields
    adjunctive_procedures: Optional[List[str]] = []
    clinical_notes: Optional[str] = None
    surgery_date: Optional[str] = None
    prosthetic_loading_date: Optional[str] = None
    implant_outcome: Optional[str] = "Pending"
    medical_red_flags: Optional[List[Dict[str, Any]]] = []
    diameter_mm: Optional[float] = None
    length_mm: Optional[float] = None
    osseointegration_success: bool = False
    peri_implant_health: bool = False
    site_specific_notes: Optional[str] = None
    complication_remarks: Optional[str] = None
    case_number: Optional[str] = None
    # NEW FIELDS from Reference Excel
    arch: Optional[str] = "Upper"  # Upper/Lower
    jaw_region: Optional[str] = "Anterior"  # Anterior/Posterior
    implant_system: Optional[str] = None  # Product line/system name
    cover_screw: bool = False
    healing_abutment: bool = False
    membrane_used: bool = False
    isq_value: Optional[float] = None  # Implant Stability Quotient
    follow_up_date: Optional[str] = None
    surgeon_name: Optional[str] = None
    consultant_surgeon: Optional[str] = None   # visiting/consultant surgeon who performed the surgery
    tag_image: Optional[str] = None   # base64 data URL of the implant package label
    # Progress tracking
    current_stage: int = 1  # 1=Placement, 2=Second Stage/Impressions, 3=Prosthesis Delivery
    osseointegration_days: int = 90  # Editable healing period in days
    stage_2_date: Optional[str] = None  # Date stage 2 was logged
    stage_3_date: Optional[str] = None  # Date stage 3 was logged

class ImplantStageUpdate(BaseModel):
    current_stage: Optional[int] = None
    osseointegration_days: Optional[int] = None
    stage_2_date: Optional[str] = None
    stage_3_date: Optional[str] = None

class FPDCreate(BaseModel):
    patient_id: str
    tooth_numbers: List[int]
    tooth_roles: Optional[Dict[str, str]] = {}   # { "13": "abutment", "14": "pontic", "15": "abutment" }
    prosthetic_loading_date: Optional[str] = None
    crown_count: str = "Single"  # Single / Multiple
    connected_implant_ids: Optional[List[str]] = []
    crown_type: str = "Screw Retained"  # Cement Retained / Screw Retained
    crown_material: str = "Zirconia"  # Metal / Porcelain fused to metal / Zirconia
    clinical_notes: Optional[str] = None
    clinic_id: Optional[str] = None
    consultant_prosthodontist: Optional[str] = None   # visiting/consultant who did the FPD case
    lab_name: Optional[str] = None                   # dental lab that fabricated the crowns
    warranty_image: Optional[str] = None             # base64 or storage path of warranty card photo

class ProfileUpdate(BaseModel):
    college: Optional[str] = None
    college_place: Optional[str] = None
    phone: Optional[str] = None
    specialization: Optional[str] = None
    currency: Optional[str] = None
    name: Optional[str] = None
    country: Optional[str] = None
    bio: Optional[str] = None

# Auth Endpoints
@api_router.post("/auth/register")
async def register(doctor: DoctorRegister, response: Response):
    email_lower = doctor.email.lower()
    existing = await db.users.find_one({"email": email_lower})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    password_hash = hash_password(doctor.password)
    user_doc = {
        "email": email_lower,
        "password_hash": password_hash,
        "name": doctor.name,
        "phone": doctor.phone,
        "country": doctor.country,
        "currency": doctor.currency,
        "registration_number": doctor.registration_number,
        "college": doctor.college,
        "specialization": doctor.specialization,
        "profile_picture": doctor.profile_picture,
        # Extended profile fields — persisted so data is complete for PostgreSQL migration
        "gender": doctor.gender,
        "date_of_birth": doctor.date_of_birth,
        "address_street": doctor.address_street,
        "address_city": doctor.address_city,
        "address_state": doctor.address_state,
        "address_zip": doctor.address_zip,
        "designation": doctor.designation,
        "organization": doctor.organization,
        "years_of_experience": doctor.years_of_experience,
        "primary_clinic": doctor.primary_clinic,
        "consulting_clinics": doctor.consulting_clinics,
        "clinical_focus": doctor.clinical_focus,
        "education": doctor.education or [],
        "publications": doctor.publications or [],
        "role": "doctor",
        "created_at": datetime.now(timezone.utc)
    }
    
    result = await db.users.insert_one(user_doc)
    user_id = str(result.inserted_id)
    
    access_token = create_access_token(user_id, email_lower)
    refresh_token = create_refresh_token(user_id)
    
    response.set_cookie(
        key="access_token",
        value=access_token,
        httponly=True,
        secure=False,
        samesite="lax",
        max_age=900,
        path="/"
    )
    response.set_cookie(
        key="refresh_token",
        value=refresh_token,
        httponly=True,
        secure=False,
        samesite="lax",
        max_age=604800,
        path="/"
    )
    
    user_doc["_id"] = user_id
    user_doc.pop("password_hash")
    return user_doc

@api_router.post("/auth/login")
async def login(credentials: DoctorLogin, response: Response):
    email_lower = credentials.email.lower()
    user = await db.users.find_one({"email": email_lower})
    
    if not user or not verify_password(credentials.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    user_id = str(user["_id"])
    access_token = create_access_token(user_id, email_lower)
    refresh_token = create_refresh_token(user_id)
    
    response.set_cookie(
        key="access_token",
        value=access_token,
        httponly=True,
        secure=False,
        samesite="lax",
        max_age=900,
        path="/"
    )
    response.set_cookie(
        key="refresh_token",
        value=refresh_token,
        httponly=True,
        secure=False,
        samesite="lax",
        max_age=604800,
        path="/"
    )
    
    user["_id"] = user_id
    user.pop("password_hash")
    return user

@api_router.get("/auth/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    return user

# Subscription plan limits (storage in MB)
PLAN_STORAGE = { "free": 500, "pro": 5120, "clinic": 20480 }

# ── External Ad Slots (editable here or via DB in future) ──
EXTERNAL_ADS = [
    {
        "id": "ad_001",
        "image_url": None,                          # None = show text-only placeholder
        "headline": "Advertise Here",
        "subtext": "Reach dental professionals using DentalHub. Contact us to place your ad.",
        "cta_label": "Learn More",
        "cta_url": None,
        "bg_color": "#F0F4FF",
        "border_color": "#BFD0F7",
        "tag": "Sponsored",
        "active": True,
    },
]

@api_router.get("/ads")
async def get_ads():
    """Return active external ads — no auth required so ads load even on public pages."""
    return [ad for ad in EXTERNAL_ADS if ad.get("active")]

@api_router.get("/subscription/status")
async def get_subscription_status(request: Request):
    user = await get_current_user(request)
    plan = user.get("plan_type", "free")
    plan_end = user.get("plan_end")

    # Calculate storage used: sum file sizes in uploads/dentalhub/
    uid = user["_id"]
    used_bytes = 0
    patient_ids = [str(p["_id"]) async for p in db.patients.find({"doctor_id": uid}, {"_id": 1})]
    implant_ids = [str(i["_id"]) async for i in db.implants.find({"doctor_id": uid}, {"_id": 1})]

    # Sum clinical photo/radiograph sizes stored in implant docs
    async for imp in db.implants.find({"doctor_id": uid}):
        for ph in imp.get("clinical_photos", []):
            used_bytes += ph.get("size", 0)
        for rg in imp.get("radiographs", []):
            used_bytes += rg.get("size", 0)

    # Sum extra patient photos
    async for ph in db.patient_photos.find({"doctor_id": uid}):
        used_bytes += ph.get("size", 0)

    used_mb = round(used_bytes / (1024 * 1024), 2)
    limit_mb = PLAN_STORAGE.get(plan, 500)

    return {
        "plan": plan,
        "plan_end": plan_end.isoformat() if hasattr(plan_end, "isoformat") else plan_end,
        "used_mb": used_mb,
        "limit_mb": limit_mb,
        "used_pct": round(min(used_mb / limit_mb * 100, 100), 1) if limit_mb else 0,
    }

@api_router.post("/subscription/upgrade")
async def upgrade_subscription(request: Request):
    """Simulate plan upgrade — in production this would go through a payment gateway."""
    user = await get_current_user(request)
    body = await request.json()
    plan = body.get("plan", "free")
    billing = body.get("billing", "monthly")  # 'monthly' | 'yearly'
    if plan not in PLAN_STORAGE:
        raise HTTPException(status_code=400, detail="Invalid plan")
    months = 1 if billing == "monthly" else 12
    plan_end = datetime.now(timezone.utc) + timedelta(days=30 * months)
    await db.users.update_one(
        {"_id": ObjectId(user["_id"])},
        {"$set": {"plan_type": plan, "plan_billing": billing, "plan_end": plan_end}}
    )
    return {"message": f"Upgraded to {plan} ({billing})", "plan_end": plan_end.isoformat()}

@api_router.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user_id = payload["sub"]
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Refresh token expired — please log in again")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid refresh token")
    access_token = create_access_token(user_id)
    response.set_cookie(
        key="access_token", value=access_token,
        httponly=True, secure=False, samesite="lax",
        max_age=900, path="/"
    )
    return {"message": "Token refreshed"}

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logged out successfully"}

# Patient Endpoints
@api_router.post("/patients")
async def create_patient(patient: PatientCreate, request: Request):
    user = await get_current_user(request)
    patient_doc = patient.model_dump()
    patient_doc["doctor_id"] = user["_id"]
    patient_doc["created_at"] = datetime.now(timezone.utc)
    
    result = await db.patients.insert_one(patient_doc)
    patient_doc["_id"] = str(result.inserted_id)
    return patient_doc

@api_router.get("/patients")
async def get_patients(request: Request):
    user = await get_current_user(request)
    patients = await db.patients.find({"doctor_id": user["_id"]}, {"password_hash": 0}).to_list(1000)
    for patient in patients:
        patient["_id"] = str(patient["_id"])
    return patients

@api_router.get("/patients/{patient_id}")
async def get_patient(patient_id: str, request: Request):
    user = await get_current_user(request)
    try:
        patient = await db.patients.find_one({"_id": ObjectId(patient_id), "doctor_id": user["_id"]})
        if not patient:
            raise HTTPException(status_code=404, detail="Patient not found")
        patient["_id"] = str(patient["_id"])
        return patient
    except:
        raise HTTPException(status_code=404, detail="Patient not found")

@api_router.patch("/patients/{patient_id}/tooth-conditions")
async def save_tooth_conditions(patient_id: str, request: Request):
    """Save the tooth condition map (missing/healthy) for a patient."""
    user = await get_current_user(request)
    try:
        obj_id = ObjectId(patient_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid patient ID")
    body = await request.json()
    conditions = body.get("tooth_conditions", {})
    result = await db.patients.update_one(
        {"_id": obj_id, "doctor_id": user["_id"]},
        {"$set": {"tooth_conditions": conditions}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Patient not found")
    return {"message": "Tooth conditions saved"}

@api_router.put("/patients/{patient_id}")
async def update_patient(patient_id: str, patient: PatientCreate, request: Request):
    user = await get_current_user(request)
    try:
        obj_id = ObjectId(patient_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid patient ID")
    existing = await db.patients.find_one({"_id": obj_id, "doctor_id": user["_id"]})
    if not existing:
        raise HTTPException(status_code=404, detail="Patient not found")
    new_data = patient.model_dump()
    # Compute which fields changed (skip profile_picture from diff)
    tracked_fields = ["name", "age", "gender", "phone", "email", "address",
                      "medical_history", "emergency_phone", "alternate_email"]
    changes = []
    for field in tracked_fields:
        old_val = existing.get(field)
        new_val = new_data.get(field)
        if str(old_val or '') != str(new_val or ''):
            changes.append({"field": field, "old": old_val, "new": new_val})
    await db.patients.update_one({"_id": obj_id}, {"$set": new_data})
    if changes:
        await db.patient_edit_logs.insert_one({
            "patient_id": patient_id,
            "doctor_id": user["_id"],
            "changed_at": datetime.now(timezone.utc),
            "changes": changes,
        })
    return {"message": "Patient updated successfully"}

@api_router.get("/patients/{patient_id}/edit-log")
async def get_patient_edit_log(patient_id: str, request: Request):
    user = await get_current_user(request)
    logs = await db.patient_edit_logs.find(
        {"patient_id": patient_id, "doctor_id": user["_id"]}
    ).sort("changed_at", -1).to_list(50)
    for log in logs:
        log["_id"] = str(log["_id"])
        log["changed_at"] = log["changed_at"].isoformat()
    return logs

@api_router.post("/patients/{patient_id}/photos")
async def upload_patient_photo(patient_id: str, file: UploadFile = File(...), caption: str = Query(''), request: Request = None):
    """Upload an extra photo directly attached to a patient (not implant-specific)."""
    user = await get_current_user(request)
    try:
        obj_id = ObjectId(patient_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid patient ID")
    patient = await db.patients.find_one({"_id": obj_id, "doctor_id": user["_id"]})
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    path = f"{APP_NAME}/patients/{patient_id}/photos/{uuid.uuid4()}.{ext}"
    data = await file.read()
    result = put_object(path, data, file.content_type or "image/jpeg")
    photo_doc = {
        "id": str(uuid.uuid4()),
        "patient_id": patient_id,
        "doctor_id": user["_id"],
        "storage_path": result["path"],
        "caption": caption,
        "original_filename": file.filename,
        "content_type": file.content_type,
        "size": result["size"],
        "uploaded_at": datetime.now(timezone.utc),
    }
    await db.patient_photos.insert_one(photo_doc)
    photo_doc["_id"] = str(photo_doc.get("_id", ""))
    photo_doc["uploaded_at"] = photo_doc["uploaded_at"].isoformat()
    return photo_doc

@api_router.get("/patients/{patient_id}/photos")
async def get_patient_photos(patient_id: str, request: Request):
    user = await get_current_user(request)
    photos = await db.patient_photos.find({"patient_id": patient_id, "doctor_id": user["_id"]}).sort("uploaded_at", -1).to_list(100)
    for p in photos:
        p["_id"] = str(p["_id"])
        if hasattr(p.get("uploaded_at"), "isoformat"):
            p["uploaded_at"] = p["uploaded_at"].isoformat()
    return photos

@api_router.delete("/patients/{patient_id}/photos/{photo_id}")
async def delete_patient_photo(patient_id: str, photo_id: str, request: Request):
    user = await get_current_user(request)
    result = await db.patient_photos.delete_one({"id": photo_id, "patient_id": patient_id, "doctor_id": user["_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Photo not found")
    return {"message": "Photo deleted"}

@api_router.post("/patients/{patient_id}/profile-picture")
async def upload_patient_profile_picture(patient_id: str, file: UploadFile = File(...), request: Request = None):
    user = await get_current_user(request)
    try:
        obj_id = ObjectId(patient_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid patient ID")
    patient = await db.patients.find_one({"_id": obj_id, "doctor_id": user["_id"]})
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    path = f"{APP_NAME}/patients/{patient_id}/profile/{uuid.uuid4()}.{ext}"
    data = await file.read()
    put_object(path, data, file.content_type or "image/jpeg")
    await db.patients.update_one({"_id": obj_id}, {"$set": {"profile_picture": path}})
    return {"profile_picture": path}

# Clinic Endpoints
@api_router.post("/clinics")
async def create_clinic(clinic: ClinicCreate, request: Request):
    user = await get_current_user(request)
    clinic_doc = clinic.model_dump()
    clinic_doc["doctor_id"] = user["_id"]
    clinic_doc["created_at"] = datetime.now(timezone.utc)
    
    result = await db.clinics.insert_one(clinic_doc)
    clinic_doc["_id"] = str(result.inserted_id)
    return clinic_doc

@api_router.get("/clinics")
async def get_clinics(request: Request):
    user = await get_current_user(request)
    clinics = await db.clinics.find({"doctor_id": user["_id"]}).to_list(100)
    for clinic in clinics:
        clinic["_id"] = str(clinic["_id"])
    return clinics

# Implant Endpoints
@api_router.post("/implants")
async def create_implant(implant: ImplantCreate, request: Request):
    user = await get_current_user(request)
    # Verify the patient exists and belongs to this doctor before creating the implant.
    # Without this check a doctor could attach clinical records to another doctor's patient.
    try:
        patient = await db.patients.find_one({"_id": ObjectId(implant.patient_id), "doctor_id": user["_id"]})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid patient ID")
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    implant_doc = implant.model_dump()
    implant_doc["doctor_id"] = user["_id"]
    implant_doc["created_at"] = datetime.now(timezone.utc)
    implant_doc["osseointegration_date"] = datetime.now(timezone.utc) + timedelta(days=90)
    implant_doc["status"] = implant_doc.get("implant_outcome", "Pending")
    
    # Generate case number if not provided
    if not implant_doc.get("case_number"):
        year = datetime.now().year
        count = await db.implants.count_documents({"doctor_id": user["_id"]}) + 1
        implant_doc["case_number"] = f"IMPLANT-{year}-{count:03d}"
    
    # Initialize photo and radiograph arrays
    implant_doc["clinical_photos"] = []
    implant_doc["radiographs"] = []
    
    result = await db.implants.insert_one(implant_doc)
    implant_doc["_id"] = str(result.inserted_id)
    return implant_doc

@api_router.get("/implants")
async def get_implants(request: Request, patient_id: Optional[str] = None):
    user = await get_current_user(request)
    query = {"doctor_id": user["_id"]}
    if patient_id:
        query["patient_id"] = patient_id
    
    implants = await db.implants.find(query).to_list(1000)
    for implant in implants:
        implant["_id"] = str(implant["_id"])
    return implants

@api_router.get("/implants/due-for-second-stage")
async def get_implants_due_for_second_stage(request: Request):
    """Return stage-1 implants whose osseointegration period has elapsed."""
    user = await get_current_user(request)
    today = datetime.now(timezone.utc)

    stage1_implants = await db.implants.find({
        "doctor_id": user["_id"],
        "$or": [{"current_stage": 1}, {"current_stage": {"$exists": False}}],
        "surgery_date": {"$exists": True, "$ne": None, "$ne": ""},
    }).to_list(1000)

    due = []
    for implant in stage1_implants:
        surgery_date_str = implant.get("surgery_date")
        if not surgery_date_str:
            continue
        try:
            surgery_date = datetime.fromisoformat(surgery_date_str)
            if surgery_date.tzinfo is None:
                surgery_date = surgery_date.replace(tzinfo=timezone.utc)
        except Exception:
            continue

        osseo_days = implant.get("osseointegration_days", 90)
        days_elapsed = (today - surgery_date).days
        if days_elapsed >= osseo_days:
            patient_name = "Unknown"
            patient_id_str = None
            try:
                patient = await db.patients.find_one({"_id": ObjectId(implant["patient_id"])})
                if patient:
                    patient_name = patient["name"]
                    patient_id_str = str(patient["_id"])
            except Exception:
                pass

            due.append({
                "implant_id": str(implant["_id"]),
                "tooth_number": implant.get("tooth_number"),
                "brand": implant.get("brand"),
                "patient_name": patient_name,
                "patient_id": patient_id_str,
                "days_elapsed": days_elapsed,
                "osseointegration_days": osseo_days,
                "surgery_date": surgery_date_str,
                "case_number": implant.get("case_number"),
            })

    return due

@api_router.get("/implants/{implant_id}")
async def get_implant(implant_id: str, request: Request):
    user = await get_current_user(request)
    try:
        implant = await db.implants.find_one({"_id": ObjectId(implant_id), "doctor_id": user["_id"]})
        if not implant:
            raise HTTPException(status_code=404, detail="Implant not found")
        implant["_id"] = str(implant["_id"])
        return implant
    except:
        raise HTTPException(status_code=404, detail="Implant not found")

@api_router.patch("/implants/{implant_id}/stage")
async def update_implant_stage(implant_id: str, update: ImplantStageUpdate, request: Request):
    """Advance treatment stage or edit osseointegration days for a single implant."""
    user = await get_current_user(request)
    # Only include fields that were explicitly sent in the request body
    patch = update.model_dump(exclude_unset=True)
    if not patch:
        raise HTTPException(status_code=400, detail="No fields to update")
    try:
        obj_id = ObjectId(implant_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid implant ID")
    result = await db.implants.update_one(
        {"_id": obj_id, "doctor_id": user["_id"]},
        {"$set": patch}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Implant not found")
    return {"message": "Stage updated successfully"}

@api_router.put("/implants/{implant_id}")
async def update_implant(implant_id: str, implant: ImplantCreate, request: Request):
    user = await get_current_user(request)
    try:
        result = await db.implants.update_one(
            {"_id": ObjectId(implant_id), "doctor_id": user["_id"]},
            {"$set": implant.model_dump()}
        )
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Implant not found")
        return {"message": "Implant updated successfully"}
    except:
        raise HTTPException(status_code=404, detail="Implant not found")

# File Upload Endpoints

# FPD Record Endpoints
@api_router.post("/fpd-records")
async def create_fpd_record(fpd: FPDCreate, request: Request):
    user = await get_current_user(request)
    # Verify the patient exists and belongs to this doctor before creating the FPD record.
    try:
        patient = await db.patients.find_one({"_id": ObjectId(fpd.patient_id), "doctor_id": user["_id"]})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid patient ID")
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    fpd_doc = fpd.model_dump()
    fpd_doc["doctor_id"] = user["_id"]
    fpd_doc["created_at"] = datetime.now(timezone.utc)

    year = datetime.now().year
    count = await db.fpd_records.count_documents({"doctor_id": user["_id"]}) + 1
    fpd_doc["case_number"] = f"FPD-{year}-{count:03d}"

    result = await db.fpd_records.insert_one(fpd_doc)
    fpd_doc["_id"] = str(result.inserted_id)
    return fpd_doc

@api_router.get("/fpd-records")
async def get_fpd_records(request: Request, patient_id: Optional[str] = None):
    user = await get_current_user(request)
    query = {"doctor_id": user["_id"]}
    if patient_id:
        query["patient_id"] = patient_id
    records = await db.fpd_records.find(query).to_list(1000)
    for r in records:
        r["_id"] = str(r["_id"])
    return records

@api_router.put("/fpd-records/{record_id}")
async def update_fpd_record(record_id: str, fpd: FPDCreate, request: Request):
    user = await get_current_user(request)
    try:
        obj_id = ObjectId(record_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid record ID")
    update_data = fpd.model_dump()
    update_data.pop("patient_id", None)  # don't overwrite patient_id
    result = await db.fpd_records.update_one(
        {"_id": obj_id, "doctor_id": user["_id"]},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="FPD record not found or not yours")
    return {"message": "FPD record updated successfully"}

@api_router.post("/fpd-records/{record_id}/warranty-image")
async def upload_fpd_warranty_image(record_id: str, file: UploadFile = File(...), request: Request = None):
    """Upload a warranty card photo for a FPD record."""
    user = await get_current_user(request)
    try:
        obj_id = ObjectId(record_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid record ID")
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "jpg"
    path = f"{APP_NAME}/fpd/{record_id}/warranty_{uuid.uuid4()}.{ext}"
    contents = await file.read()
    put_object(path, contents, file.content_type or "image/jpeg")
    await db.fpd_records.update_one(
        {"_id": obj_id, "doctor_id": user["_id"]},
        {"$set": {"warranty_image": path}}
    )
    return {"warranty_image": path}

@api_router.delete("/fpd-records/{record_id}")
async def delete_fpd_record(record_id: str, request: Request):
    user = await get_current_user(request)
    result = await db.fpd_records.delete_one({"_id": ObjectId(record_id), "doctor_id": user["_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="FPD record not found")
    return {"message": "FPD record deleted"}

# Profile Update Endpoint
@api_router.get("/public/profile/{doctor_id}")
async def get_public_profile(doctor_id: str):
    """Public doctor profile — no auth required. Used for shareable link & QR code."""
    try:
        obj_id = ObjectId(doctor_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid doctor ID")
    user = await db.users.find_one({"_id": obj_id})
    if not user:
        raise HTTPException(status_code=404, detail="Doctor not found")

    uid = str(user["_id"])

    # Stats
    patient_count = await db.patients.count_documents({"doctor_id": uid})
    implants = await db.implants.find({"doctor_id": uid}).to_list(10000)
    total_implants = len(implants)
    successful = sum(1 for i in implants if i.get("osseointegration_success") is True)
    success_rate = round((successful / total_implants) * 100) if total_implants > 0 else None
    fpd_count = await db.fpd_records.count_documents({"doctor_id": uid})

    # Years active
    created_at = user.get("created_at")
    years_active = None
    if created_at:
        from datetime import datetime, timezone
        years_active = max(1, (datetime.now(timezone.utc) - created_at).days // 365)

    return {
        "id": uid,
        "name": user.get("name", ""),
        "specialization": user.get("specialization", ""),
        "registration_number": user.get("registration_number", ""),
        "college": user.get("college", ""),
        "college_place": user.get("college_place", ""),
        "country": user.get("country", ""),
        "bio": user.get("bio", ""),
        "profile_picture": user.get("profile_picture"),
        "clinics": user.get("clinics", []),
        "stats": {
            "patients": patient_count,
            "implants": total_implants,
            "fpd": fpd_count,
            "success_rate": success_rate,
            "years_active": years_active,
        }
    }

@api_router.put("/auth/profile")
async def update_profile(profile: ProfileUpdate, request: Request):
    user = await get_current_user(request)
    update_data = {k: v for k, v in profile.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    await db.users.update_one({"_id": ObjectId(user["_id"])}, {"$set": update_data})
    updated = await db.users.find_one({"_id": ObjectId(user["_id"])})
    updated["_id"] = str(updated["_id"])
    updated.pop("password_hash", None)
    return updated

@api_router.post("/implants/{implant_id}/photos")
async def upload_photo(
    implant_id: str,
    file: UploadFile = File(...),
    view_type: str = Query(...),
    request: Request = None
):
    user = await get_current_user(request)
    
    # Verify implant belongs to doctor
    implant = await db.implants.find_one({"_id": ObjectId(implant_id), "doctor_id": user["_id"]})
    if not implant:
        raise HTTPException(status_code=404, detail="Implant not found")
    
    # Upload to storage
    ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    path = f"{APP_NAME}/implants/{implant_id}/photos/{uuid.uuid4()}.{ext}"
    data = await file.read()
    result = put_object(path, data, file.content_type or "image/jpeg")
    
    # Store reference in database
    photo_doc = {
        "id": str(uuid.uuid4()),
        "implant_id": implant_id,
        "storage_path": result["path"],
        "view_type": view_type,
        "original_filename": file.filename,
        "content_type": file.content_type,
        "size": result["size"],
        "uploaded_at": datetime.now(timezone.utc)
    }
    
    await db.implants.update_one(
        {"_id": ObjectId(implant_id)},
        {"$push": {"clinical_photos": photo_doc}}
    )
    
    return photo_doc

@api_router.post("/implants/{implant_id}/radiographs")
async def upload_radiograph(
    implant_id: str,
    file: UploadFile = File(...),
    view_type: str = Query(...),
    request: Request = None
):
    user = await get_current_user(request)
    
    # Verify implant belongs to doctor
    implant = await db.implants.find_one({"_id": ObjectId(implant_id), "doctor_id": user["_id"]})
    if not implant:
        raise HTTPException(status_code=404, detail="Implant not found")
    
    # Upload to storage
    ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    path = f"{APP_NAME}/implants/{implant_id}/radiographs/{uuid.uuid4()}.{ext}"
    data = await file.read()
    result = put_object(path, data, file.content_type or "image/jpeg")
    
    # Store reference in database
    radiograph_doc = {
        "id": str(uuid.uuid4()),
        "implant_id": implant_id,
        "storage_path": result["path"],
        "view_type": view_type,
        "original_filename": file.filename,
        "content_type": file.content_type,
        "size": result["size"],
        "uploaded_at": datetime.now(timezone.utc)
    }
    
    await db.implants.update_one(
        {"_id": ObjectId(implant_id)},
        {"$push": {"radiographs": radiograph_doc}}
    )
    
    return radiograph_doc

@api_router.get("/files/{file_path:path}")
async def download_file(
    file_path: str,
    authorization: str = Header(None),
    auth: str = Query(None),
    request: Request = None
):
    # Support cookie auth (for <img src> requests), Bearer header, or query param
    cookie_token = request.cookies.get("access_token") if request else None
    if not authorization and not auth and not cookie_token:
        raise HTTPException(status_code=401, detail="Not authenticated")

    try:
        # Get file from storage
        data, content_type = get_object(file_path)
        return Response(content=data, media_type=content_type)
    except Exception as e:
        raise HTTPException(status_code=404, detail="File not found")

# Analytics Endpoints
@api_router.get("/analytics/overview")
async def get_analytics_overview(request: Request):
    user = await get_current_user(request)
    
    total_patients = await db.patients.count_documents({"doctor_id": user["_id"]})
    total_implants = await db.implants.count_documents({"doctor_id": user["_id"]})
    total_clinics = await db.clinics.count_documents({"doctor_id": user["_id"]})
    
    # Implants by type
    implant_types = await db.implants.aggregate([
        {"$match": {"doctor_id": user["_id"]}},
        {"$group": {"_id": "$implant_type", "count": {"$sum": 1}}}
    ]).to_list(100)
    
    # Pending osseointegration
    pending_osseo = await db.implants.count_documents({
        "doctor_id": user["_id"],
        "status": "healing",
        "osseointegration_date": {"$gte": datetime.now(timezone.utc)}
    })
    
    return {
        "total_patients": total_patients,
        "total_implants": total_implants,
        "total_clinics": total_clinics,
        "implant_types": implant_types,
        "pending_osseointegration": pending_osseo
    }

@api_router.get("/analytics/brand-surgeon-summary")
async def get_brand_surgeon_summary(request: Request):
    user = await get_current_user(request)
    
    # Brand analysis
    brand_analysis = await db.implants.aggregate([
        {"$match": {"doctor_id": user["_id"]}},
        {"$group": {
            "_id": "$brand",
            "count": {"$sum": 1},
            "systems": {"$addToSet": "$implant_system"},
            "avg_torque": {"$avg": "$insertion_torque"},
            "avg_isq": {"$avg": "$isq_value"}
        }},
        {"$sort": {"count": -1}}
    ]).to_list(100)
    
    total_implants = await db.implants.count_documents({"doctor_id": user["_id"]})
    
    # Add percentage
    for brand in brand_analysis:
        brand["percentage"] = round((brand["count"] / total_implants * 100), 2) if total_implants > 0 else 0
        brand["systems"] = [s for s in brand["systems"] if s]  # Filter out None
    
    # Surgeon analysis
    surgeon_analysis = await db.implants.aggregate([
        {"$match": {"doctor_id": user["_id"]}},
        {"$group": {
            "_id": "$surgeon_name",
            "implants_placed": {"$sum": 1},
            "sites_used": {"$addToSet": "$tooth_number"},
            "brands_used": {"$addToSet": "$brand"}
        }},
        {"$sort": {"implants_placed": -1}}
    ]).to_list(100)
    
    for surgeon in surgeon_analysis:
        surgeon["percentage"] = round((surgeon["implants_placed"] / total_implants * 100), 2) if total_implants > 0 else 0
        surgeon["unique_sites"] = len(surgeon["sites_used"])
        surgeon["brands_used"] = [b for b in surgeon["brands_used"] if b]
    
    return {
        "brands": brand_analysis,
        "surgeons": surgeon_analysis,
        "total_implants": total_implants
    }

@api_router.get("/analytics/site-usage")
async def get_site_usage_analysis(request: Request):
    user = await get_current_user(request)
    
    # Site analysis with arch and jaw region
    site_analysis = await db.implants.aggregate([
        {"$match": {"doctor_id": user["_id"]}},
        {"$group": {
            "_id": {
                "tooth_number": "$tooth_number",
                "arch": "$arch",
                "jaw_region": "$jaw_region"
            },
            "count": {"$sum": 1},
            "last_surgery": {"$max": "$surgery_date"},
            "brands": {"$addToSet": "$brand"}
        }},
        {"$sort": {"count": -1}}
    ]).to_list(100)
    
    total_implants = await db.implants.count_documents({"doctor_id": user["_id"]})
    
    for site in site_analysis:
        site["percentage"] = round((site["count"] / total_implants * 100), 2) if total_implants > 0 else 0
        site["tooth_number"] = site["_id"]["tooth_number"]
        site["arch"] = site["_id"]["arch"]
        site["jaw_region"] = site["_id"]["jaw_region"]
        site["brands"] = [b for b in site["brands"] if b]
        del site["_id"]
    
    return {
        "sites": site_analysis,
        "total_implants": total_implants
    }

@api_router.get("/analytics/dimension-analysis")
async def get_dimension_analysis(request: Request):
    user = await get_current_user(request)
    
    # Diameter analysis
    diameter_analysis = await db.implants.aggregate([
        {"$match": {"doctor_id": user["_id"], "diameter_mm": {"$ne": None}}},
        {"$group": {
            "_id": "$diameter_mm",
            "count": {"$sum": 1}
        }},
        {"$sort": {"_id": 1}}
    ]).to_list(100)
    
    # Length analysis
    length_analysis = await db.implants.aggregate([
        {"$match": {"doctor_id": user["_id"], "length_mm": {"$ne": None}}},
        {"$group": {
            "_id": "$length_mm",
            "count": {"$sum": 1}
        }},
        {"$sort": {"_id": 1}}
    ]).to_list(100)
    
    # Combination matrix
    combination_matrix = await db.implants.aggregate([
        {"$match": {"doctor_id": user["_id"], "diameter_mm": {"$ne": None}, "length_mm": {"$ne": None}}},
        {"$group": {
            "_id": {
                "diameter": "$diameter_mm",
                "length": "$length_mm"
            },
            "count": {"$sum": 1}
        }}
    ]).to_list(500)
    
    total_implants = await db.implants.count_documents({"doctor_id": user["_id"]})
    
    # Add percentages
    for diam in diameter_analysis:
        diam["percentage"] = round((diam["count"] / total_implants * 100), 2) if total_implants > 0 else 0
        diam["diameter"] = diam["_id"]
        del diam["_id"]
    
    for length in length_analysis:
        length["percentage"] = round((length["count"] / total_implants * 100), 2) if total_implants > 0 else 0
        length["length"] = length["_id"]
        del length["_id"]
    
    # Format matrix for easy consumption
    matrix = {}
    for combo in combination_matrix:
        diam = combo["_id"]["diameter"]
        length = combo["_id"]["length"]
        if diam not in matrix:
            matrix[diam] = {}
        matrix[diam][length] = combo["count"]
    
    return {
        "diameters": diameter_analysis,
        "lengths": length_analysis,
        "matrix": matrix,
        "total_implants": total_implants
    }

@api_router.get("/analytics/financial")
async def get_financial_analytics(request: Request):
    user = await get_current_user(request)
    
    # Mock financial data - in production, you'd track costs and revenue
    implants = await db.implants.find({"doctor_id": user["_id"]}).to_list(1000)
    
    # Sample cost per implant type
    costs = {
        "single": 1500,
        "bridge": 4500,
        "full_mouth": 25000
    }
    
    total_revenue = 0
    for implant in implants:
        implant_type = implant.get("implant_type", "single").lower()
        total_revenue += costs.get(implant_type, 1500)
    
    return {
        "total_revenue": total_revenue,
        "total_implants": len(implants),
        "average_per_implant": total_revenue / len(implants) if len(implants) > 0 else 0
    }

# ── AI CHAT ───────────────────────────────────────────────────────────────────

class ChatMessage(BaseModel):
    role: str
    content: str

class ChatRequest(BaseModel):
    messages: List[ChatMessage]
    patient_id: Optional[str] = None  # context: which patient page the user is on

CHAT_TOOLS = [
    {
        "name": "create_patient",
        "description": "Create a new patient record. Use when the user asks to add a new patient.",
        "input_schema": {
            "type": "object",
            "properties": {
                "name":            {"type": "string",  "description": "Full name"},
                "age":             {"type": "integer", "description": "Age in years"},
                "gender":          {"type": "string",  "enum": ["Male", "Female", "Other"]},
                "phone":           {"type": "string",  "description": "Phone number (optional)"},
                "medical_history": {"type": "string",  "description": "Medical conditions, allergies (optional)"},
            },
            "required": ["name", "age", "gender"],
        },
    },
    {
        "name": "log_implant",
        "description": "Log a dental implant for an existing patient. Use when the user describes placing an implant.",
        "input_schema": {
            "type": "object",
            "properties": {
                "patient_id":       {"type": "string",  "description": "Patient ID (use context patient_id if on patient page)"},
                "patient_name":     {"type": "string",  "description": "Patient name to search if no ID"},
                "tooth_number":     {"type": "integer", "description": "FDI tooth number (11-48)"},
                "brand":            {"type": "string",  "description": "Implant brand e.g. Nobel, Straumann, Osstem"},
                "diameter_mm":      {"type": "number",  "description": "Diameter in mm"},
                "length_mm":        {"type": "number",  "description": "Length in mm"},
                "insertion_torque": {"type": "number",  "description": "Insertion torque in Ncm (optional)"},
                "surgery_date":     {"type": "string",  "description": "Surgery date YYYY-MM-DD, default today"},
                "surgeon_name":     {"type": "string",  "description": "Surgeon name (optional)"},
                "notes":            {"type": "string",  "description": "Clinical notes (optional)"},
            },
            "required": ["tooth_number", "brand"],
        },
    },
    {
        "name": "get_patient_summary",
        "description": "Get a summary of a patient's implants, FPD records and history.",
        "input_schema": {
            "type": "object",
            "properties": {
                "patient_id":   {"type": "string", "description": "Patient ID"},
                "patient_name": {"type": "string", "description": "Patient name to search if no ID"},
            },
        },
    },
    {
        "name": "answer_clinical_question",
        "description": "Answer a general dental implantology clinical question using knowledge. No database lookup needed.",
        "input_schema": {
            "type": "object",
            "properties": {
                "question": {"type": "string", "description": "The clinical question to answer"},
            },
            "required": ["question"],
        },
    },
]

async def _resolve_patient(db, doctor_id: str, patient_id: Optional[str], patient_name: Optional[str]):
    if patient_id:
        try:
            p = await db.patients.find_one({"_id": ObjectId(patient_id), "doctor_id": doctor_id})
            if p:
                return p
        except Exception:
            pass
    if patient_name:
        p = await db.patients.find_one({
            "doctor_id": doctor_id,
            "name": {"$regex": patient_name, "$options": "i"},
            "deleted": {"$ne": True},
        })
        return p
    return None

@api_router.post("/chat")
async def chat(req: ChatRequest, request: Request):
    user = await get_current_user(request)
    doctor_id = user["_id"]

    anthropic_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not anthropic_key:
        return {
            "role": "assistant",
            "content": "⚠️ AI chat is not configured yet. Ask your administrator to add the ANTHROPIC_API_KEY to the server environment.",
            "action": None,
        }

    try:
        import anthropic as _anthropic
    except ImportError:
        return {
            "role": "assistant",
            "content": "⚠️ The `anthropic` Python package is not installed on the server. Run: pip install anthropic",
            "action": None,
        }

    client_ai = _anthropic.Anthropic(api_key=anthropic_key)

    system_prompt = (
        "You are OSIOLOG Assistant, an AI embedded in a dental implant case management system. "
        "You help dentists quickly add patient data, log implants, and look up clinical information — all by chat. "
        "Be concise and clinical. Always confirm what action you took. "
        "Today's date is " + datetime.now().strftime("%Y-%m-%d") + ". "
        "When logging an implant and no patient_id is given, use the patient_name tool input to search. "
        "If the user is on a patient page, patient_id context is provided in the system prompt below."
    )
    if req.patient_id:
        system_prompt += f"\n\nCurrent page context: patient_id = {req.patient_id}."

    messages = [{"role": m.role, "content": m.content} for m in req.messages]

    response = client_ai.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1024,
        system=system_prompt,
        tools=CHAT_TOOLS,
        messages=messages,
    )

    action_result = None

    # Handle tool use
    if response.stop_reason == "tool_use":
        tool_block = next((b for b in response.content if b.type == "tool_use"), None)
        if tool_block:
            tool_name = tool_block.name
            args = tool_block.input

            # ── create_patient ──
            if tool_name == "create_patient":
                new_patient = {
                    "doctor_id": doctor_id,
                    "name": args["name"],
                    "age": args.get("age"),
                    "gender": args.get("gender", "Other"),
                    "phone": args.get("phone", ""),
                    "medical_history": args.get("medical_history", ""),
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "deleted": False,
                }
                result = await db.patients.insert_one(new_patient)
                patient_id_new = str(result.inserted_id)
                action_result = {"type": "patient_created", "patient_id": patient_id_new, "name": args["name"]}
                tool_result_text = f"Patient '{args['name']}' created successfully. ID: {patient_id_new}"

            # ── log_implant ──
            elif tool_name == "log_implant":
                pid = args.get("patient_id") or req.patient_id
                patient = await _resolve_patient(db, doctor_id, pid, args.get("patient_name"))
                if not patient:
                    tool_result_text = f"Could not find patient '{args.get('patient_name', '')}'. Please create the patient first."
                else:
                    tn = args["tooth_number"]
                    arch = "Upper" if tn <= 28 else "Lower"
                    tens = tn // 10
                    jaw_region = "Anterior" if (tens in [1,2,3,4] and tn % 10 <= 3) else "Posterior"
                    today_str = datetime.now().strftime("%Y-%m-%d")
                    implant_doc = {
                        "doctor_id": doctor_id,
                        "patient_id": str(patient["_id"]),
                        "tooth_number": tn,
                        "brand": args["brand"],
                        "diameter_mm": args.get("diameter_mm"),
                        "length_mm": args.get("length_mm"),
                        "insertion_torque": args.get("insertion_torque"),
                        "surgery_date": args.get("surgery_date", today_str),
                        "surgeon_name": args.get("surgeon_name", ""),
                        "clinical_notes": args.get("notes", ""),
                        "arch": arch,
                        "jaw_region": jaw_region,
                        "implant_outcome": "Pending",
                        "created_at": datetime.now(timezone.utc).isoformat(),
                    }
                    res = await db.implants.insert_one(implant_doc)
                    action_result = {"type": "implant_logged", "implant_id": str(res.inserted_id), "patient_id": str(patient["_id"]), "tooth": tn}
                    size_str = ""
                    if args.get("diameter_mm") and args.get("length_mm"):
                        size_str = f" {args['diameter_mm']}×{args['length_mm']}mm"
                    tool_result_text = f"Implant logged: tooth {tn},{size_str} {args['brand']} for patient {patient['name']}."

            # ── get_patient_summary ──
            elif tool_name == "get_patient_summary":
                pid = args.get("patient_id") or req.patient_id
                patient = await _resolve_patient(db, doctor_id, pid, args.get("patient_name"))
                if not patient:
                    tool_result_text = "Patient not found."
                else:
                    implants = await db.implants.find({"patient_id": str(patient["_id"])}).to_list(100)
                    fpds = await db.fpd_logs.find({"patient_id": str(patient["_id"])}).to_list(50)
                    summary_lines = [f"Patient: {patient['name']}, {patient.get('age','?')}y {patient.get('gender','')}"]
                    if patient.get("medical_history"):
                        summary_lines.append(f"Medical history: {patient['medical_history']}")
                    summary_lines.append(f"Implants: {len(implants)}")
                    for imp in implants:
                        summary_lines.append(f"  • Tooth {imp.get('tooth_number')} — {imp.get('brand','')} {imp.get('diameter_mm','')}×{imp.get('length_mm','')}mm, {imp.get('surgery_date','')}")
                    if fpds:
                        summary_lines.append(f"FPD records: {len(fpds)}")
                    tool_result_text = "\n".join(summary_lines)

            # ── answer_clinical_question ──
            elif tool_name == "answer_clinical_question":
                tool_result_text = f"[Answering clinical question: {args['question']}]"
            else:
                tool_result_text = "Unknown tool."

            # Send tool result back to Claude for final response
            followup_messages = messages + [
                {"role": "assistant", "content": response.content},
                {"role": "user",      "content": [{"type": "tool_result", "tool_use_id": tool_block.id, "content": tool_result_text}]},
            ]
            final = client_ai.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=512,
                system=system_prompt,
                tools=CHAT_TOOLS,
                messages=followup_messages,
            )
            reply_text = next((b.text for b in final.content if hasattr(b, "text")), "Done.")
        else:
            reply_text = next((b.text for b in response.content if hasattr(b, "text")), "")
    else:
        reply_text = next((b.text for b in response.content if hasattr(b, "text")), "")

    return {"role": "assistant", "content": reply_text, "action": action_result}

# ── BACKUP / RESTORE ──────────────────────────────────────────────────────────

@api_router.get("/backup/export")
async def export_backup(request: Request):
    """Export all doctor-scoped data as a JSON backup."""
    user = await get_current_user(request)
    uid = user["_id"]

    patients   = await db.patients.find({"doctor_id": uid}).to_list(10000)
    implants   = await db.implants.find({"doctor_id": uid}).to_list(10000)
    fpd_records = await db.fpd_records.find({"doctor_id": uid}).to_list(10000)
    clinics    = await db.clinics.find({"doctor_id": uid}).to_list(10000)
    edit_logs  = await db.patient_edit_logs.find({"doctor_id": uid}).to_list(10000)

    def clean(docs):
        out = []
        for d in docs:
            d["_id"] = str(d["_id"])
            # Stringify any remaining ObjectId / datetime fields
            for k, v in d.items():
                if hasattr(v, 'isoformat'):
                    d[k] = v.isoformat()
            out.append(d)
        return out

    backup = {
        "version": "1.0",
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "doctor": {
            "email": user.get("email"),
            "name":  user.get("name"),
        },
        "patients":    clean(patients),
        "implants":    clean(implants),
        "fpd_records": clean(fpd_records),
        "clinics":     clean(clinics),
        "edit_logs":   clean(edit_logs),
    }
    return backup


@api_router.post("/backup/restore")
async def restore_backup(request: Request):
    """Restore from a JSON backup. Merges data — does NOT delete existing records."""
    user = await get_current_user(request)
    uid = user["_id"]

    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON payload")

    version = payload.get("version")
    if not version:
        raise HTTPException(status_code=400, detail="Unrecognised backup format")

    stats = {"patients": 0, "implants": 0, "fpd_records": 0, "clinics": 0}

    async def restore_collection(col_name, items, key_field="name"):
        count = 0
        for item in items:
            item.pop("_id", None)
            item["doctor_id"] = uid          # reassign to current doctor
            # Upsert by a natural key to avoid duplicates on repeated restore
            existing = await db[col_name].find_one({
                "doctor_id": uid,
                key_field: item.get(key_field),
            })
            if not existing:
                await db[col_name].insert_one(item)
                count += 1
        return count

    stats["patients"]    = await restore_collection("patients",    payload.get("patients", []),    "name")
    stats["implants"]    = await restore_collection("implants",    payload.get("implants", []),    "tooth_number")
    stats["fpd_records"] = await restore_collection("fpd_records", payload.get("fpd_records", []), "prosthetic_loading_date")
    stats["clinics"]     = await restore_collection("clinics",     payload.get("clinics", []),     "name")

    return {"message": "Restore complete", "inserted": stats}


@api_router.post("/bulk-import")
async def bulk_import(file: UploadFile = File(...), request: Request = None):
    """
    Accept an Excel workbook with sheets: Patients, Implants, FPD.
    Creates records in DB, linking implants and FPD to patients by patient name.
    Returns counts of created records and any row-level errors.
    """
    import openpyxl, io
    user = await get_current_user(request)
    uid = user["_id"]

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        wb_formulas = openpyxl.load_workbook(io.BytesIO(contents), data_only=False)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    # Accept alternate sheet names (reference doc uses different names)
    SHEET_ALIASES = {
        "Patients":  ["Patients", "Patient Log", "Patients Log", "Patient Data"],
        "Implants":  ["Implants", "Implant Log", "Implant Data", "Patient Log"],
        "FPD":       ["FPD", "FPD Log", "Crown", "Crowns"],
    }

    # Column name aliases — map reference-doc column names → our canonical names
    COLUMN_ALIASES = {
        "Patient Name":            ["Patient Name", "patient name", "Name", "Patient"],
        "Age":                     ["Age", "age"],
        "Gender":                  ["Gender", "gender", "Sex"],
        "Phone":                   ["Phone", "phone", "Mobile", "Contact"],
        "Medical History":         ["Medical History", "medical history", "Medical"],
        "Tooth Number":            ["Tooth Number", "Tooth / Site #", "Tooth", "Site", "Tooth No", "tooth_number", "Tooth#"],
        "Brand":                   ["Brand", "Implant Brand", "brand", "Manufacturer"],
        "Implant System":          ["Implant System", "System", "implant_system", "System Name"],
        "Size (Diameter mm)":      ["Size (Diameter mm)", "Implant Diameter (mm)", "Diameter", "Diameter (mm)", "diameter_mm", "Size"],
        "Length (mm)":             ["Length (mm)", "Implant Length (mm)", "Length", "length_mm"],
        "Insertion Torque (Ncm)":  ["Insertion Torque (Ncm)", "Torque Value (Ncm)", "Torque", "Insertion Torque", "torque"],
        "Cover Screw":             ["Cover Screw", "cover_screw"],
        "Healing Abutment":        ["Healing Abutment", "healing_abutment"],
        "Bone Graft":              ["Bone Graft", "Bone Graft Used", "bone_graft", "Graft"],
        "Membrane Used":           ["Membrane Used", "membrane_used", "Membrane"],
        "ISQ Value":               ["ISQ Value", "ISQ", "isq_value"],
        "Surgery Date":            ["Surgery Date", "Date of Surgery", "Surgery", "surgery_date"],
        "Follow Up Date":          ["Follow Up Date", "Follow-up Date", "Followup Date", "follow_up_date"],
        "Surgeon Name":            ["Surgeon Name", "Surgeon", "surgeon_name", "Doctor"],
        "Arch":                    ["Arch", "arch"],
        "Jaw Region":              ["Jaw Region", "jaw_region", "Region"],
        "Clinical Notes":          ["Clinical Notes", "Notes / Complications", "Notes", "clinical_notes", "Complications"],
    }

    def normalise_headers(raw_headers):
        """Map any column name variant to our canonical name."""
        result = []
        alias_map = {}
        for canon, variants in COLUMN_ALIASES.items():
            for v in variants:
                alias_map[v.lower().strip()] = canon
        for h in raw_headers:
            if h is None:
                result.append("")
                continue
            clean = str(h).split("←")[0].split("<-")[0].strip()
            result.append(alias_map.get(clean.lower(), clean))
        return result

    def find_sheet(canonical_name):
        """Return the first matching sheet name from aliases, or None."""
        for alias in SHEET_ALIASES.get(canonical_name, [canonical_name]):
            if alias in wb.sheetnames:
                return alias
        return None

    def sheet_rows(sheet_name):
        actual = find_sheet(sheet_name)
        if not actual:
            return []
        ws = wb[actual]
        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not all_rows:
            return []

        # Find the actual header row — skip title/subtitle rows (merged cells, no column structure)
        # A header row has multiple non-null values across columns
        header_row_idx = 0
        for i, row in enumerate(all_rows[:5]):
            non_null = [v for v in row if v is not None]
            if len(non_null) >= 3:
                header_row_idx = i
                break

        raw_headers = all_rows[header_row_idx]
        headers = normalise_headers(raw_headers)

        # Data starts one row after the header row
        # But also skip any notes/instruction row immediately after headers
        data_start = header_row_idx + 1
        if data_start < len(all_rows):
            next_row = all_rows[data_start]
            non_null = [v for v in next_row if v is not None]
            # Notes row: all strings, no numbers
            if non_null and all(isinstance(v, str) for v in non_null):
                data_start += 1

        actual_formula = find_sheet(sheet_name)
        ws_formula = wb_formulas[actual_formula] if actual_formula and actual_formula in wb_formulas.sheetnames else None
        pat_actual = find_sheet("Patients")
        pat_ws_formula = wb_formulas[pat_actual] if pat_actual and pat_actual in wb_formulas.sheetnames else None

        def resolve_patient_name_from_formula(excel_row_1based):
            """If col A of this row is a formula referencing Patients sheet, resolve it."""
            if ws_formula is None or pat_ws_formula is None:
                return None
            cell = ws_formula.cell(row=excel_row_1based, column=1)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                # Extract row number from formula like =IF(Patients!A5="","",Patients!A5)
                import re
                m = re.search(r'Patients!A(\d+)', cell.value)
                if m:
                    src_row = int(m.group(1))
                    src_cell = pat_ws_formula.cell(row=src_row, column=1)
                    if src_cell.value:
                        return str(src_cell.value).strip()
            return None

        rows = []
        for idx, row in enumerate(all_rows[data_start:]):
            if all(v is None for v in row):
                continue
            row_dict = dict(zip(headers, row))
            # Fix patient name if formula cell returned None
            if not row_dict.get("Patient Name") or str(row_dict.get("Patient Name", "")).strip() in ("", "0", "None"):
                excel_row = data_start + idx + 1  # 1-based row in sheet
                resolved = resolve_patient_name_from_formula(excel_row)
                if resolved:
                    row_dict["Patient Name"] = resolved
            rows.append(row_dict)
        return rows

    def val(row, *keys, default=None):
        for k in keys:
            # Try exact key and also "Patient Name" when header has been normalised
            v = row.get(k)
            if v is not None and str(v).strip() != "" and str(v).strip() != "0":
                return str(v).strip()
        return default

    def boolval(row, key):
        v = val(row, key, default="")
        return str(v).lower() in ("yes", "true", "1", "y")

    def parse_date(row, *keys):
        """Accept DD-MM-YYYY, YYYY-MM-DD, or Excel serial number → YYYY-MM-DD string."""
        raw = val(row, *keys)
        if not raw:
            return None
        raw_str = str(raw).strip()
        # Excel serial date (e.g. 45306 = a number with no dashes)
        try:
            serial = float(raw_str)
            if serial > 1000:  # looks like a serial, not a day/month number
                from datetime import date as _date
                excel_epoch = _date(1899, 12, 30)
                from datetime import timedelta as _td
                return (_date.fromordinal(excel_epoch.toordinal() + int(serial))).strftime("%Y-%m-%d")
        except (ValueError, OverflowError):
            pass
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(raw_str, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return raw_str  # return as-is if no format matched

    errors = []
    results = {"patients": 0, "implants": 0, "fpd": 0}

    # ── PATIENTS ──
    # If no dedicated Patients sheet, auto-extract unique patient names from Implants sheet
    patient_name_map = {}  # name (lower) → _id string
    pat_rows = sheet_rows("Patients")
    if not pat_rows:
        # Reference-doc style: extract unique names from Patient Log / Implants sheet
        seen_names = {}
        for row in sheet_rows("Implants"):
            name = val(row, "Patient Name")
            if name and name.lower() not in seen_names:
                seen_names[name.lower()] = {"Patient Name": name, "Age": "", "Gender": "", "Phone": "", "Email": "", "Address": "", "Medical History": ""}
        pat_rows = list(seen_names.values())

    for i, row in enumerate(pat_rows, start=2):
        name = val(row, "Patient Name", "name")
        if not name:
            errors.append(f"Patients row {i}: Missing patient name — skipped")
            continue
        try:
            age = int(float(val(row, "Age", "age") or 0))
        except Exception:
            age = 0
        doc = {
            "doctor_id": uid,
            "name": name,
            "age": age,
            "gender": val(row, "Gender", "gender") or "Other",
            "phone": val(row, "Phone", "phone") or "",
            "email": val(row, "Email", "email"),
            "alternate_email": val(row, "Alternate Email", "alternate_email"),
            "emergency_phone": val(row, "Emergency Phone", "emergency_phone"),
            "address": val(row, "Address", "address"),
            "medical_history": val(row, "Medical History", "medical_history"),
            "created_at": datetime.now(timezone.utc),
        }
        result = await db.patients.insert_one(doc)
        pid = str(result.inserted_id)
        patient_name_map[name.lower()] = pid
        results["patients"] += 1

    # Also load existing patients of this doctor into the map (for re-uploads)
    async for p in db.patients.find({"doctor_id": uid}, {"_id": 1, "name": 1}):
        key = p["name"].lower()
        if key not in patient_name_map:
            patient_name_map[key] = str(p["_id"])

    # ── BUILD CLINIC MAP (name → _id), create new clinics on the fly ──
    clinic_name_map = {}  # clinic name (lower) → _id string
    async for c in db.clinics.find({"doctor_id": uid}, {"_id": 1, "name": 1}):
        clinic_name_map[c["name"].lower()] = str(c["_id"])

    async def resolve_clinic(clinic_name, clinic_address=""):
        if not clinic_name:
            return None
        key = clinic_name.strip().lower()
        if key in clinic_name_map:
            return clinic_name_map[key]
        # Create new clinic
        doc = {
            "doctor_id": uid,
            "name": clinic_name.strip(),
            "address": clinic_address.strip() if clinic_address else "",
            "created_at": datetime.now(timezone.utc),
        }
        result = await db.clinics.insert_one(doc)
        cid = str(result.inserted_id)
        clinic_name_map[key] = cid
        return cid

    # ── IMPLANTS ──
    last_pname = None  # carry forward patient name when col A is blank (multiple implants per patient)
    for i, row in enumerate(sheet_rows("Implants"), start=2):
        pname = val(row, "Patient Name", "patient_name")
        if pname:
            last_pname = pname
        else:
            pname = last_pname
        if not pname:
            errors.append(f"Implants row {i}: Missing patient name — skipped")
            continue
        pid = patient_name_map.get(pname.lower())
        if not pid:
            errors.append(f"Implants row {i}: Patient '{pname}' not found — skipped")
            continue
        try:
            raw_tooth = val(row, "Tooth Number", "tooth_number") or "0"
            raw_tooth = str(raw_tooth).replace("#", "").strip()  # handle "#14" → "14"
            tooth = int(float(raw_tooth))
        except Exception:
            errors.append(f"Implants row {i}: Invalid tooth number — skipped")
            continue
        doc = {
            "doctor_id": uid,
            "patient_id": pid,
            "tooth_number": tooth,
            "implant_type": val(row, "Implant Type") or "Single",
            "brand": val(row, "Brand") or "",
            "size": val(row, "Size (Diameter mm)", "size") or "",
            "length": val(row, "Length (mm)", "length") or "",
            "diameter_mm": float(val(row, "Size (Diameter mm)", "diameter_mm") or 0) or None,
            "length_mm": float(val(row, "Length (mm)", "length_mm") or 0) or None,
            "insertion_torque": float(val(row, "Insertion Torque (Ncm)", "insertion_torque") or 0) or None,
            "connection_type": val(row, "Connection Type", "connection_type") or "",
            "surgical_approach": val(row, "Surgical Approach", "surgical_approach") or "Flapless",
            "arch": val(row, "Arch") or "Upper",
            "jaw_region": val(row, "Jaw Region") or "Anterior",
            "implant_system": val(row, "Implant System"),
            "bone_graft": val(row, "Bone Graft"),
            "sinus_lift_type": val(row, "Sinus Lift Type"),
            "is_pterygoid": boolval(row, "Pterygoid"),
            "is_zygomatic": boolval(row, "Zygomatic"),
            "is_subperiosteal": boolval(row, "Subperiosteal"),
            "cover_screw": boolval(row, "Cover Screw"),
            "healing_abutment": boolval(row, "Healing Abutment"),
            "membrane_used": boolval(row, "Membrane Used"),
            "isq_value": float(val(row, "ISQ Value", "isq_value") or 0) or None,
            "surgery_date": parse_date(row, "Surgery Date", "surgery_date"),
            "prosthetic_loading_date": parse_date(row, "Prosthetic Loading Date", "prosthetic_loading_date"),
            "follow_up_date": parse_date(row, "Follow Up Date", "follow_up_date"),
            "surgeon_name": val(row, "Surgeon Name", "surgeon_name"),
            "implant_outcome": val(row, "Outcome", "implant_outcome") or "Pending",
            "osseointegration_success": boolval(row, "Osseointegration Success"),
            "peri_implant_health": boolval(row, "Peri-Implant Health"),
            "notes": val(row, "Notes"),
            "clinical_notes": val(row, "Clinical Notes"),
            "case_number": val(row, "Case Number"),
            "consultant_surgeon": val(row, "Consultant Surgeon"),
            "clinic_id": await resolve_clinic(val(row, "Clinic Name"), val(row, "Clinic Address")),
            "current_stage": 1,
            "osseointegration_days": 90,
            "clinical_photos": [],
            "radiographs": [],
            "created_at": datetime.now(timezone.utc),
        }
        result = await db.implants.insert_one(doc)
        results["implants"] += 1

    # ── FPD ──
    last_fpd_pname = None
    for i, row in enumerate(sheet_rows("FPD"), start=2):
        pname = val(row, "Patient Name", "patient_name")
        if pname:
            last_fpd_pname = pname
        else:
            pname = last_fpd_pname
        if not pname:
            errors.append(f"FPD row {i}: Missing patient name — skipped")
            continue
        pid = patient_name_map.get(pname.lower())
        if not pid:
            errors.append(f"FPD row {i}: Patient '{pname}' not found — skipped")
            continue
        teeth_raw = val(row, "Tooth Numbers", "tooth_numbers") or ""
        try:
            tooth_numbers = [int(t.strip()) for t in str(teeth_raw).split(",") if t.strip().isdigit()]
        except Exception:
            tooth_numbers = []
        if not tooth_numbers:
            errors.append(f"FPD row {i}: No valid tooth numbers — skipped")
            continue
        doc = {
            "doctor_id": uid,
            "patient_id": pid,
            "tooth_numbers": tooth_numbers,
            "crown_count": val(row, "Crown Count") or "Single",
            "crown_type": val(row, "Crown Type") or "Screw Retained",
            "crown_material": val(row, "Crown Material") or "Zirconia",
            "prosthetic_loading_date": parse_date(row, "Prosthetic Loading Date"),
            "consultant_prosthodontist": val(row, "Consultant Prosthodontist"),
            "lab_name": val(row, "Dental Lab"),
            "clinical_notes": val(row, "Clinical Notes"),
            "connected_implant_ids": [],
            "created_at": datetime.now(timezone.utc),
        }
        await db.fpd_records.insert_one(doc)
        results["fpd"] += 1

    return {"message": "Import complete", "created": results, "errors": errors}


@api_router.post("/bulk-import-sheets")
async def bulk_import_from_sheets(request: Request):
    """
    Import implant data directly from a public Google Sheet.
    Accepts a Google Sheets URL, fetches the 'Patient Log' sheet as CSV,
    and creates patients + implants in the DB.
    Supports same-name consecutive rows = multiple implants for same patient.
    """
    import csv, io, urllib.request
    user = await get_current_user(request)
    uid = user["_id"]

    body = await request.json()
    sheet_url = (body.get("sheet_url") or "").strip()
    if not sheet_url:
        raise HTTPException(status_code=400, detail="sheet_url is required")

    # Extract the sheet ID from any Google Sheets URL format
    import re
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", sheet_url)
    if not m:
        raise HTTPException(status_code=400, detail="Invalid Google Sheets URL. Copy the URL from your browser's address bar.")
    sheet_id = m.group(1)

    # Fetch the first sheet (gid=0) as CSV — works for any publicly shared sheet
    csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid=0"
    try:
        with urllib.request.urlopen(csv_url, timeout=15) as resp:
            raw = resp.read().decode("utf-8-sig")
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail="Could not read the Google Sheet. Make sure it is shared as 'Anyone with the link can view'."
        )

    reader = csv.reader(io.StringIO(raw))
    all_rows = list(reader)
    if not all_rows:
        raise HTTPException(status_code=400, detail="The sheet appears to be empty.")

    # Find header row — first row with 3+ non-empty values
    header_idx = 0
    for i, row in enumerate(all_rows[:5]):
        if sum(1 for c in row if c.strip()) >= 3:
            header_idx = i
            break

    raw_headers = all_rows[header_idx]

    # Column alias map — matches reference sheet column names
    COLUMN_ALIASES = {
        "Patient Name":           ["Patient Name", "Name", "Patient"],
        "Date of Surgery":        ["Date of Surgery", "Surgery Date", "Date"],
        "Surgeon":                ["Surgeon", "Surgeon Name", "Doctor", "Operator"],
        "Tooth / Site #":         ["Tooth / Site #", "Tooth Number", "Tooth", "Site", "Tooth#", "Tooth No"],
        "Arch":                   ["Arch"],
        "Jaw Region":             ["Jaw Region", "Region"],
        "Implant Brand":          ["Implant Brand", "Brand", "Manufacturer"],
        "Implant System":         ["Implant System", "System"],
        "Implant Diameter (mm)":  ["Implant Diameter (mm)", "Diameter (mm)", "Diameter", "Size (Diameter mm)", "Size"],
        "Implant Length (mm)":    ["Implant Length (mm)", "Length (mm)", "Length"],
        "Cover Screw":            ["Cover Screw"],
        "Healing Abutment":       ["Healing Abutment"],
        "Bone Graft Used":        ["Bone Graft Used", "Bone Graft", "Graft"],
        "Membrane Used":          ["Membrane Used", "Membrane"],
        "Torque Value (Ncm)":     ["Torque Value (Ncm)", "Insertion Torque (Ncm)", "Torque"],
        "ISQ Value":              ["ISQ Value", "ISQ"],
        "Follow-up Date":         ["Follow-up Date", "Follow Up Date", "Followup Date"],
        "Notes / Complications":  ["Notes / Complications", "Notes", "Clinical Notes", "Complications"],
        "Clinic Name":            ["Clinic Name", "Clinic", "Hospital Name", "Hospital"],
        "Clinic Address":         ["Clinic Address", "Hospital Address", "Address"],
    }
    alias_map = {}
    for canon, variants in COLUMN_ALIASES.items():
        for v in variants:
            alias_map[v.lower().strip()] = canon

    def norm(h):
        return alias_map.get(h.lower().strip(), h.strip())

    headers = [norm(h) for h in raw_headers]

    # Skip header row and any notes/instruction row immediately after
    data_rows = all_rows[header_idx + 1:]
    if data_rows and all(not c.strip().replace(",", "").replace(".", "").isdigit() for c in data_rows[0] if c.strip()):
        # If first data row looks like all-text (instructions row), skip it
        first = [c for c in data_rows[0] if c.strip()]
        if first and all(not any(ch.isdigit() for ch in c) for c in first):
            data_rows = data_rows[1:]

    def row_dict(row):
        return {headers[i]: row[i].strip() if i < len(row) else "" for i in range(len(headers))}

    def val(d, *keys):
        for k in keys:
            v = d.get(k, "").strip()
            if v and v not in ("0", "-", "N/A", "n/a"):
                return v
        return None

    def boolval(d, key):
        return (d.get(key, "") or "").lower().strip() in ("yes", "true", "1", "y")

    def parse_date(s):
        if not s:
            return None
        s = s.strip()
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return s or None

    errors = []
    results = {"patients": 0, "implants": 0, "fpd": 0}
    patient_name_map = {}  # name_lower → patient_id
    clinic_name_map  = {}  # name_lower → clinic_id

    # Load existing patients for this doctor to avoid duplicates
    async for p in db.patients.find({"doctor_id": uid}, {"_id": 1, "name": 1}):
        patient_name_map[p["name"].lower().strip()] = str(p["_id"])

    # Load existing clinics
    async for c in db.clinics.find({"doctor_id": uid}, {"_id": 1, "name": 1}):
        clinic_name_map[c["name"].lower().strip()] = str(c["_id"])

    async def resolve_clinic_gs(clinic_name, clinic_address=""):
        if not clinic_name:
            return None
        key = clinic_name.strip().lower()
        if key in clinic_name_map:
            return clinic_name_map[key]
        doc = {
            "doctor_id": uid,
            "name": clinic_name.strip(),
            "address": (clinic_address or "").strip(),
            "created_at": datetime.now(timezone.utc),
        }
        res = await db.clinics.insert_one(doc)
        cid = str(res.inserted_id)
        clinic_name_map[key] = cid
        return cid

    last_pname = None

    for i, raw_row in enumerate(data_rows, start=header_idx + 2):
        if not any(c.strip() for c in raw_row):
            continue  # skip fully blank rows

        rd = row_dict(raw_row)

        # Patient name carry-forward: blank name = same patient as previous row
        pname = val(rd, "Patient Name")
        if pname:
            last_pname = pname
        else:
            pname = last_pname
        if not pname:
            errors.append(f"Row {i}: No patient name — skipped")
            continue

        pname_key = pname.lower().strip()

        # Create patient if not seen yet
        if pname_key not in patient_name_map:
            pdoc = {
                "doctor_id": uid,
                "name": pname,
                "age": 0,
                "gender": "Other",
                "phone": "",
                "email": None,
                "address": "",
                "medical_history": "",
                "created_at": datetime.now(timezone.utc),
            }
            res = await db.patients.insert_one(pdoc)
            patient_name_map[pname_key] = str(res.inserted_id)
            results["patients"] += 1

        pid = patient_name_map[pname_key]

        # Parse tooth number
        raw_tooth = val(rd, "Tooth / Site #") or ""
        raw_tooth = raw_tooth.replace("#", "").strip()
        try:
            tooth = int(float(raw_tooth))
        except Exception:
            errors.append(f"Row {i}: Invalid tooth number '{raw_tooth}' — skipped")
            continue

        # Arch / Jaw from sheet (already filled in the reference doc)
        arch = val(rd, "Arch") or ("Upper" if tooth < 30 else "Lower")
        jaw = val(rd, "Jaw Region") or ("Anterior" if tooth % 10 <= 3 else "Posterior")

        diameter_raw = val(rd, "Implant Diameter (mm)")
        length_raw   = val(rd, "Implant Length (mm)")
        torque_raw   = val(rd, "Torque Value (Ncm)")
        isq_raw      = val(rd, "ISQ Value")

        try: diameter = float(diameter_raw) if diameter_raw else None
        except: diameter = None
        try: length = float(length_raw) if length_raw else None
        except: length = None
        try: torque = float(torque_raw) if torque_raw else None
        except: torque = None
        try: isq = float(isq_raw) if isq_raw else None
        except: isq = None

        surgeon = val(rd, "Surgeon")
        surgery_date = parse_date(val(rd, "Date of Surgery"))
        follow_up    = parse_date(val(rd, "Follow-up Date"))

        clinic_id = await resolve_clinic_gs(val(rd, "Clinic Name"), val(rd, "Clinic Address"))

        implant_doc = {
            "doctor_id": uid,
            "patient_id": pid,
            "tooth_number": tooth,
            "brand": val(rd, "Implant Brand") or "",
            "implant_system": val(rd, "Implant System"),
            "size": str(diameter) if diameter else "",
            "length": str(length) if length else "",
            "diameter_mm": diameter,
            "length_mm": length,
            "insertion_torque": torque,
            "arch": arch,
            "jaw_region": jaw,
            "cover_screw": boolval(rd, "Cover Screw"),
            "healing_abutment": boolval(rd, "Healing Abutment"),
            "bone_graft": val(rd, "Bone Graft Used"),
            "membrane_used": boolval(rd, "Membrane Used"),
            "isq_value": isq,
            "surgery_date": surgery_date,
            "follow_up_date": follow_up,
            "surgeon_name": surgeon,
            "clinical_notes": val(rd, "Notes / Complications"),
            "clinic_id": clinic_id,
            "implant_type": "Single",
            "connection_type": "",
            "surgical_approach": "Flapless",
            "implant_outcome": "Pending",
            "osseointegration_success": False,
            "peri_implant_health": False,
            "current_stage": 1,
            "osseointegration_days": 90,
            "clinical_photos": [],
            "radiographs": [],
            "created_at": datetime.now(timezone.utc),
        }
        await db.implants.insert_one(implant_doc)
        results["implants"] += 1

    return {"message": "Import complete", "created": results, "errors": errors}


app.include_router(api_router)

# Startup event
@app.on_event("startup")
async def startup_event():
    logger.info(f"Local file storage ready at: {UPLOAD_DIR}")

    # Create indexes
    await db.users.create_index("email", unique=True)
    await db.patients.create_index("doctor_id")
    await db.implants.create_index("doctor_id")
    await db.implants.create_index("patient_id")
    await db.clinics.create_index("doctor_id")
    
    # Seed admin
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@dentalapp.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    
    if existing is None:
        hashed = hash_password(admin_password)
        await db.users.insert_one({
            "email": admin_email,
            "password_hash": hashed,
            "name": "Admin User",
            "phone": "+1234567890",
            "country": "USA",
            "registration_number": "ADMIN001",
            "college": "Admin College",
            "specialization": "General Dentistry",
            "role": "admin",
            "created_at": datetime.now(timezone.utc)
        })
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one(
            {"email": admin_email},
            {"$set": {"password_hash": hash_password(admin_password)}}
        )
    
    # Create demo doctor and patients
    demo_doctor_email = "doctor@dentalapp.com"
    demo_doctor = await db.users.find_one({"email": demo_doctor_email})
    
    if not demo_doctor:
        demo_doctor_pass = hash_password("doctor123")
        demo_result = await db.users.insert_one({
            "email": demo_doctor_email,
            "password_hash": demo_doctor_pass,
            "name": "Dr. Sarah Johnson",
            "phone": "+1555123456",
            "country": "USA",
            "registration_number": "DDS12345",
            "college": "Harvard School of Dental Medicine",
            "specialization": "Implantology",
            "role": "doctor",
            "created_at": datetime.now(timezone.utc)
        })
        demo_doctor_id = str(demo_result.inserted_id)
        
        # Create demo clinic
        clinic_result = await db.clinics.insert_one({
            "name": "Downtown Dental Clinic",
            "address": "123 Main St, New York, NY 10001",
            "phone": "+1555999888",
            "email": "info@downtowndental.com",
            "doctor_id": demo_doctor_id,
            "created_at": datetime.now(timezone.utc)
        })
        clinic_id = str(clinic_result.inserted_id)
        
        # Create demo patients
        patient1_result = await db.patients.insert_one({
            "name": "John Smith",
            "age": 45,
            "gender": "Male",
            "phone": "+1555111222",
            "email": "john.smith@email.com",
            "address": "456 Oak Ave, New York, NY",
            "medical_history": "Hypertension, controlled with medication",
            "doctor_id": demo_doctor_id,
            "created_at": datetime.now(timezone.utc)
        })
        
        patient2_result = await db.patients.insert_one({
            "name": "Emma Williams",
            "age": 38,
            "gender": "Female",
            "phone": "+1555333444",
            "email": "emma.w@email.com",
            "address": "789 Pine St, New York, NY",
            "medical_history": "No significant medical history",
            "doctor_id": demo_doctor_id,
            "created_at": datetime.now(timezone.utc)
        })
        
        # Create demo implants
        await db.implants.insert_one({
            "patient_id": str(patient1_result.inserted_id),
            "tooth_number": 14,
            "implant_type": "Single",
            "brand": "Straumann",
            "size": "4.1mm",
            "length": "10mm",
            "insertion_torque": 35.0,
            "connection_type": "Internal Hex",
            "surgical_approach": "Immediate Placement",
            "bone_graft": "Xenograft",
            "sinus_lift_type": None,
            "is_pterygoid": False,
            "is_zygomatic": False,
            "is_subperiosteal": False,
            "clinic_id": clinic_id,
            "doctor_id": demo_doctor_id,
            "created_at": datetime.now(timezone.utc) - timedelta(days=30),
            "osseointegration_date": datetime.now(timezone.utc) + timedelta(days=60),
            "status": "healing",
            "notes": "Primary stability achieved, no complications"
        })
        
        await db.implants.insert_one({
            "patient_id": str(patient2_result.inserted_id),
            "tooth_number": 26,
            "implant_type": "Single",
            "brand": "Nobel Biocare",
            "size": "4.3mm",
            "length": "13mm",
            "insertion_torque": 40.0,
            "connection_type": "Conical",
            "surgical_approach": "Delayed Placement",
            "bone_graft": None,
            "sinus_lift_type": "Indirect",
            "is_pterygoid": False,
            "is_zygomatic": False,
            "is_subperiosteal": False,
            "clinic_id": clinic_id,
            "doctor_id": demo_doctor_id,
            "created_at": datetime.now(timezone.utc) - timedelta(days=15),
            "osseointegration_date": datetime.now(timezone.utc) + timedelta(days=75),
            "status": "healing",
            "notes": "Sinus lift performed successfully"
        })
    
    logger.info("Startup complete. Demo accounts: admin@dentalapp.com / admin123, doctor@dentalapp.com / doctor123")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()